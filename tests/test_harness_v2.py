"""v2 Phase 1 Step 4 — harness_v2 tests.

Covers ingest idempotency, view shapes, JSONL round-trip, XLSX export,
cost computation, mode.txt resolution, run-dir discovery, and the
two-distinct-latency columns required by Step 3 outcome finding 6.
"""
from __future__ import annotations

import json
import os
import tempfile
import unittest
from datetime import datetime, timezone, timedelta
from pathlib import Path
from unittest import mock

from openpyxl import load_workbook

from tools import harness_v2
from anvil import events


_FIX_ROOT = Path(__file__).resolve().parent.parent / "tools" / "fixtures" / "v2-phase-1"
# v2 Phase 2 Step 1: fixture dirs are mode-suffixed (`-mock`, `-real`),
# and the events.jsonl run_id field carries the same mode segment. The
# T1-mock + T1-real pair lives under one task_id (T1) and gives the
# composite-key regression test direct fixture coverage.
_CLEAN_DIR = _FIX_ROOT / "T1-doc-edit-mock"
_CLEAN_DIR_REAL = _FIX_ROOT / "T1-doc-edit-real"
_ESC_DIR = _FIX_ROOT / "T3-out-of-scope-real"
# v2 Phase 2 Step 2: T2-two-step-real fixture synthesises the
# "validation.fail → retry → escalate" episode shape (2 stage_b.api_end
# events with tokens, 1 validation.fail, 1 escalate). Direct fixture
# coverage for validation_failure_episodes.
_RETRY_ESC_DIR = _FIX_ROOT / "T2-two-step-real"


def _event_row(t_ms, kind, data, *, run_id, step_idx=None):
    t = datetime(2026, 5, 20, 14, 30, 0, tzinfo=timezone.utc) + timedelta(milliseconds=t_ms)
    return {
        "ts": t.isoformat(timespec="milliseconds"),
        "run_id": run_id,
        "step_idx": step_idx,
        "kind": kind,
        "data": data,
        "elapsed_ms": t_ms,
    }


class _HarnessTestBase(unittest.TestCase):

    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.tmp_path = Path(self._tmp.name)
        self.db_path = self.tmp_path / "test.duckdb"
        self.con = harness_v2.open_db(self.db_path)

    def tearDown(self) -> None:
        self.con.close()
        self._tmp.cleanup()


class TestIngestRoundTrip(_HarnessTestBase):

    def test_ingest_round_trip(self) -> None:
        # Create a synthetic run-dir under tmp.
        run_dir = self.tmp_path / "T9-synth"
        run_dir.mkdir()
        events = [
            _event_row(0, "run.start", {}, run_id="T9-synth"),
            _event_row(
                10, "planner.stage_a.api_end",
                {"model": "claude-opus-4-7", "input_tokens": 1000,
                 "output_tokens": 50, "cache_creation_input_tokens": 0,
                 "cache_read_input_tokens": 0, "duration_ms": 1500, "ok": True},
                run_id="T9-synth", step_idx=0,
            ),
            _event_row(20, "run.end", {"drops": 0}, run_id="T9-synth"),
        ]
        (run_dir / "events.jsonl").write_text(
            "\n".join(json.dumps(e) for e in events) + "\n", encoding="utf-8",
        )
        (run_dir / "mode.txt").write_text("mock\n", encoding="utf-8")
        summary = harness_v2.ingest(self.con, run_dir)
        self.assertEqual(summary["events_ingested"], 3)
        self.assertEqual(summary["task_id"], "T9")
        self.assertEqual(summary["task_label"], "synth")
        self.assertEqual(summary["mode"], "mock")
        # Round-trip via SQL.
        rows = self.con.execute(
            "SELECT kind FROM events WHERE run_id = 'T9-synth' ORDER BY ts"
        ).fetchall()
        self.assertEqual([r[0] for r in rows],
                         ["run.start", "planner.stage_a.api_end", "run.end"])

    def test_idempotent_reingest(self) -> None:
        harness_v2.ingest(self.con, _CLEAN_DIR)
        first_count = self.con.execute(
            "SELECT COUNT(*) FROM events WHERE run_id = 'T1-doc-edit-mock'"
        ).fetchone()[0]
        # Re-ingest the same dir.
        harness_v2.ingest(self.con, _CLEAN_DIR)
        second_count = self.con.execute(
            "SELECT COUNT(*) FROM events WHERE run_id = 'T1-doc-edit-mock'"
        ).fetchone()[0]
        self.assertEqual(first_count, second_count)


class TestRoutingObservabilityColumns(_HarnessTestBase):
    """v3 Phase 0 Step 1 (V3P0-1): the operations view exposes the five
    routing columns, correctly typed, with features_seen as queryable
    JSON (not a stringified blob)."""

    def _ingest_routing_run(self) -> None:
        run_dir = self.tmp_path / "T9-routing"
        run_dir.mkdir()
        routing = {
            "route_candidate": "claude-opus-4-7",
            "route_actual": "claude-opus-4-7",
            "route_fallback_fired": False,
            "policy_version": "v3-phase-0-passive",
            "features_seen": {
                "observed_prompt_token_count": 620,
                "step_idx": 0,
                "stage": "B",
                "context_paths_count": 4,
            },
        }
        events = [
            _event_row(0, "run.start", {}, run_id="T9-routing"),
            _event_row(
                10, "planner.stage_b.api_end",
                {"model": "claude-opus-4-7", "input_tokens": 620,
                 "output_tokens": 40, "cache_creation_input_tokens": 0,
                 "cache_read_input_tokens": 0, "duration_ms": 900, "ok": True,
                 **routing},
                run_id="T9-routing", step_idx=0,
            ),
            _event_row(20, "run.end", {"drops": 0}, run_id="T9-routing"),
        ]
        (run_dir / "events.jsonl").write_text(
            "\n".join(json.dumps(e) for e in events) + "\n", encoding="utf-8",
        )
        (run_dir / "mode.txt").write_text("mock\n", encoding="utf-8")
        harness_v2.ingest(self.con, run_dir)

    def test_five_columns_present_in_tuple(self) -> None:
        for col in ("route_candidate", "route_actual", "route_fallback_fired",
                    "policy_version", "features_seen"):
            self.assertIn(col, harness_v2._OPERATIONS_COLUMNS)

    def test_columns_queryable_and_typed(self) -> None:
        self._ingest_routing_run()
        row = self.con.execute(
            "SELECT route_candidate, route_actual, route_fallback_fired, "
            "       policy_version "
            "FROM operations WHERE run_id = 'T9-routing' "
            "  AND operation_kind = 'planner.stage_b.api_end'"
        ).fetchone()
        self.assertEqual(row[0], "claude-opus-4-7")
        self.assertEqual(row[1], "claude-opus-4-7")
        # route_fallback_fired comes back as a real BOOLEAN, not a string.
        self.assertIs(row[2], False)
        self.assertEqual(row[3], "v3-phase-0-passive")

    def test_features_seen_is_json_not_stringified(self) -> None:
        self._ingest_routing_run()
        # If features_seen is stored as JSON, json_extract_string can reach
        # into it directly from the view column. A stringified blob would
        # require an extra parse step and this would return NULL.
        row = self.con.execute(
            "SELECT json_extract_string(features_seen, '$.stage'), "
            "       CAST(json_extract(features_seen, '$.context_paths_count') AS BIGINT) "
            "FROM operations WHERE run_id = 'T9-routing' "
            "  AND operation_kind = 'planner.stage_b.api_end'"
        ).fetchone()
        self.assertEqual(row[0], "B")
        self.assertEqual(row[1], 4)

    def test_routing_columns_null_on_non_model_kinds(self) -> None:
        # Additive + back-compatible: a non-model-call kind (run.end isn't
        # in the operations view, so use a clean fixture's smoke.end row)
        # has NULL routing columns. Ingest the clean fixture and check.
        harness_v2.ingest(self.con, _CLEAN_DIR)
        rows = self.con.execute(
            "SELECT route_actual, policy_version FROM operations "
            "WHERE run_id = 'T1-doc-edit-mock' AND operation_kind = 'smoke.end'"
        ).fetchall()
        self.assertEqual(len(rows), 1)
        self.assertIsNone(rows[0][0])
        self.assertIsNone(rows[0][1])


class TestOperationsView(_HarnessTestBase):

    def test_clean_fixture_operations_count_and_columns(self) -> None:
        harness_v2.ingest(self.con, _CLEAN_DIR)
        rows = harness_v2.query_operations(self.con, run_id="T1-doc-edit-mock")
        # 7 cost-bearing operations for a clean single-step run:
        # planner.stage_a.api_end, planner.stage_b.api_end,
        # planner.validation.pass, coder.subprocess.end,
        # coder.scope_verify, smoke.end, git.commit.end
        self.assertEqual(len(rows), 7)
        # Column count matches the declared OPERATIONS_COLUMNS tuple.
        self.assertEqual(len(rows[0]), len(harness_v2._OPERATIONS_COLUMNS))

    def test_escalation_fixture_surfaces_escalation_reason(self) -> None:
        harness_v2.ingest(self.con, _ESC_DIR)
        rows = self.con.execute(
            "SELECT operation_kind, escalation_reason, escalation_user_latency_ms "
            "FROM operations WHERE run_id = 'T3-out-of-scope-real' "
            "  AND operation_kind = 'escalation.resolved'"
        ).fetchall()
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][0], "escalation.resolved")
        self.assertEqual(rows[0][2], 5000)
        # The escalation.resolved itself doesn't carry a reason (the reason
        # lives on escalation.raised). The operations view doesn't join
        # raised→resolved, so this column is None for the resolved row.
        # That's a documented limitation; recorded as a Step 5 carry.

    def test_cost_computation_correctness(self) -> None:
        harness_v2.ingest(self.con, _CLEAN_DIR)
        # Stage A (model=claude-opus-4-7): input=10500, output=115, cache=0/0
        #   v3 Phase 1c Step 3.5: Opus 4.7 rates $5/$25 (was Opus-4.1 $15/$75).
        #   cost = (10500*5 + 115*25) / 1e6 = (52500 + 2875) / 1e6 = 0.055375
        rows = self.con.execute(
            "SELECT cost_usd FROM operations "
            "WHERE run_id = 'T1-doc-edit-mock' AND operation_kind = 'planner.stage_a.api_end'"
        ).fetchall()
        self.assertEqual(len(rows), 1)
        self.assertAlmostEqual(rows[0][0], 0.055375, places=4)

    def test_two_distinct_latency_columns_surface(self) -> None:
        # poll_reply_chars vs escalation_user_latency_ms — they live in
        # the operations view as two distinct columns, populated only on
        # their respective kinds. Step 3 outcome finding 6.
        harness_v2.ingest(self.con, _ESC_DIR)
        cols = harness_v2._OPERATIONS_COLUMNS
        self.assertIn("escalation_user_latency_ms", cols)
        self.assertIn("poll_reply_chars", cols)
        # The escalation row has latency_ms populated, poll_reply_chars NULL.
        # No poll.reply event in fixtures → that column is NULL for all rows.
        rows = self.con.execute(
            "SELECT escalation_user_latency_ms, poll_reply_chars "
            "FROM operations WHERE run_id = 'T3-out-of-scope-real' "
            "  AND operation_kind = 'escalation.resolved'"
        ).fetchall()
        self.assertEqual(rows[0][0], 5000)
        self.assertIsNone(rows[0][1])


class TestPerRunSummary(_HarnessTestBase):

    def test_clean_fixture_summary_fields(self) -> None:
        harness_v2.ingest(self.con, _CLEAN_DIR)
        rows = harness_v2.query_per_run_summary(self.con, run_id="T1-doc-edit-mock")
        self.assertEqual(len(rows), 1)
        # Columns: run_id, task_id, task_label, mode,
        # total_cost_usd, total_duration_s, planner_calls, coder_calls,
        # escalations, resumed, terminal_event
        row = rows[0]
        # v2 Phase 2 Step 1: run_id carries `-mock` suffix; task_label
        # stays mode-independent (`doc-edit`, not `doc-edit-mock`)
        # because derive_task() strips the mode suffix before parsing.
        self.assertEqual(row[0], "T1-doc-edit-mock")
        self.assertEqual(row[1], "T1")
        self.assertEqual(row[2], "doc-edit")
        self.assertEqual(row[3], "mock")
        # 2 planner calls (stage A + stage B), 1 coder call, 0 escalations,
        # not resumed.
        self.assertEqual(row[6], 2)  # planner_calls
        self.assertEqual(row[7], 1)  # coder_calls
        self.assertEqual(row[8], 0)  # escalations
        self.assertEqual(row[9], False)  # resumed

    def test_resumed_run_flag_true(self) -> None:
        run_dir = self.tmp_path / "T9-resumed"
        run_dir.mkdir()
        events = [
            _event_row(0,  "run.start",  {}, run_id="T9-resumed"),
            _event_row(5,  "run.resume", {"run_id": "T9-resumed", "from_step": 2}, run_id="T9-resumed"),
            _event_row(10, "run.end",    {"drops": 0}, run_id="T9-resumed"),
        ]
        (run_dir / "events.jsonl").write_text(
            "\n".join(json.dumps(e) for e in events) + "\n", encoding="utf-8",
        )
        harness_v2.ingest(self.con, run_dir)
        rows = harness_v2.query_per_run_summary(self.con, run_id="T9-resumed")
        self.assertEqual(rows[0][9], True)


class TestPerTaskComparison(_HarnessTestBase):

    def test_comparison_pivots_mock_real(self) -> None:
        harness_v2.ingest(self.con, _CLEAN_DIR)        # T1 / mock
        harness_v2.ingest(self.con, _ESC_DIR)          # T3 / real
        rows = self.con.execute(
            "SELECT task_id, planner_calls_mock, planner_calls_real, "
            "       framework_overhead_s "
            "FROM per_task_comparison ORDER BY task_id"
        ).fetchall()
        # Two rows (T1, T3); each populates one side of the mock/real split.
        self.assertEqual(len(rows), 2)
        t1 = next(r for r in rows if r[0] == "T1")
        t3 = next(r for r in rows if r[0] == "T3")
        # T1 was mocked: planner_calls_mock populated, planner_calls_real None
        self.assertIsNotNone(t1[1])
        self.assertIsNone(t1[2])
        # T3 was real: planner_calls_real populated, planner_calls_mock None
        self.assertIsNone(t3[1])
        self.assertIsNotNone(t3[2])
        # framework_overhead_s == total_duration_mock (per brief).
        self.assertIsNotNone(t1[3])  # T1 has mock data → has overhead

    # v2 Phase 2 Step 1 ----------------------------------------------------
    # The composite (run_id, mode) idempotency key is the load-bearing
    # invariant of Step 1. The two tests below give it direct fixture
    # coverage: ingesting mock-then-real for the same task_id must
    # preserve both halves on disk, in the events table, and in the
    # derived per_task_comparison view.

    def test_mock_then_real_same_task_preserves_both_halves(self) -> None:
        """Mock ingest followed by real ingest for the SAME task_id must
        not clobber the mock rows. Under v2 Phase 1's run_id-only
        idempotency key this would have lost the mock half (V2P1-4)."""
        harness_v2.ingest(self.con, _CLEAN_DIR)         # T1 / mock
        mock_count = self.con.execute(
            "SELECT COUNT(*) FROM events "
            "WHERE run_id = 'T1-doc-edit-mock' AND mode = 'mock'"
        ).fetchone()[0]
        self.assertGreater(mock_count, 0)

        harness_v2.ingest(self.con, _CLEAN_DIR_REAL)    # T1 / real
        # After the real ingest, the mock rows must still be there.
        mock_count_after = self.con.execute(
            "SELECT COUNT(*) FROM events "
            "WHERE run_id = 'T1-doc-edit-mock' AND mode = 'mock'"
        ).fetchone()[0]
        real_count_after = self.con.execute(
            "SELECT COUNT(*) FROM events "
            "WHERE run_id = 'T1-doc-edit-real' AND mode = 'real'"
        ).fetchone()[0]
        self.assertEqual(mock_count_after, mock_count,
                         "real ingest clobbered the mock half")
        self.assertGreater(real_count_after, 0,
                           "real ingest produced no rows")
        # run_metadata also holds both halves.
        rm_rows = self.con.execute(
            "SELECT run_id, mode FROM run_metadata WHERE task_id = 'T1' "
            "ORDER BY mode"
        ).fetchall()
        self.assertEqual(rm_rows,
                         [("T1-doc-edit-mock", "mock"),
                          ("T1-doc-edit-real", "real")])

    def test_per_task_comparison_non_null_both_halves(self) -> None:
        """After ingesting a mock+real pair for the same task, the T1
        per_task_comparison row has BOTH mock and real columns populated
        — the load-bearing exam-question of v2 Phase 2 Q1."""
        harness_v2.ingest(self.con, _CLEAN_DIR)         # T1 / mock
        harness_v2.ingest(self.con, _CLEAN_DIR_REAL)    # T1 / real
        row = self.con.execute(
            "SELECT task_id, planner_calls_mock, planner_calls_real, "
            "       total_duration_mock, total_duration_real, "
            "       framework_overhead_s "
            "FROM per_task_comparison WHERE task_id = 'T1'"
        ).fetchone()
        self.assertIsNotNone(row)
        self.assertEqual(row[0], "T1")
        self.assertIsNotNone(row[1], "planner_calls_mock NULL")
        self.assertIsNotNone(row[2], "planner_calls_real NULL")
        self.assertIsNotNone(row[3], "total_duration_mock NULL")
        self.assertIsNotNone(row[4], "total_duration_real NULL")
        self.assertIsNotNone(row[5], "framework_overhead_s NULL")


class TestValidationFailureEpisodes(_HarnessTestBase):
    """v2 Phase 2 Step 2: validation_failure_episodes view.

    The T2-two-step-real fixture encodes the production
    "validation.fail → retry → escalate" episode shape (planner.py
    `_run_stage_b_with_retry`, second-attempt empty path). The view
    must surface one episode row for the fixture's step_idx=0 with
    n_validation_fails=1 and extra_api_calls=1 (the retry's
    stage_b.api_end is "beyond the first")."""

    def test_episode_row_shape(self) -> None:
        harness_v2.ingest(self.con, _RETRY_ESC_DIR)
        rows = harness_v2.query_validation_failure_episodes(self.con)
        self.assertEqual(len(rows), 1, f"expected 1 episode, got {rows}")
        # Columns per _VALIDATION_FAILURE_EPISODES_COLUMNS:
        #   (run_id, task_id, mode, step_idx, n_validation_fails,
        #    first_error, second_error, recovered,
        #    extra_api_calls, total_extra_cost_usd)
        r = rows[0]
        self.assertEqual(r[0], "T2-two-step-real")
        self.assertEqual(r[1], "T2")
        self.assertEqual(r[2], "real")
        self.assertEqual(r[3], 0)
        self.assertEqual(r[4], 1, "n_validation_fails")
        self.assertEqual(r[5], "missing field: scope_boundaries",
                         "first_error")
        self.assertIsNone(r[6], "second_error should be NULL")
        self.assertFalse(r[7], "recovered should be FALSE (escalated)")
        self.assertEqual(r[8], 1, "extra_api_calls")
        # total_extra_cost_usd = cost of the retry stage_b.api_end.
        # v3 Phase 1c Step 3.5: Opus 4.7 rates $5/$25 (was $15/$75).
        # (2000 input * 5 + 20 output * 25) / 1e6 = 0.0105
        self.assertAlmostEqual(r[9], 0.0105, places=6,
                               msg="total_extra_cost_usd")

    def test_empty_db_returns_no_episodes(self) -> None:
        """No validation.fail events → empty view."""
        harness_v2.ingest(self.con, _CLEAN_DIR)  # T1-mock: no fails
        rows = harness_v2.query_validation_failure_episodes(self.con)
        self.assertEqual(rows, [])


class TestModeResolution(_HarnessTestBase):

    def test_mode_present(self) -> None:
        run_dir = self.tmp_path / "T9-mode"
        run_dir.mkdir()
        (run_dir / "events.jsonl").write_text(
            json.dumps(_event_row(0, "run.start", {}, run_id="T9-mode")) + "\n",
            encoding="utf-8",
        )
        (run_dir / "mode.txt").write_text("real\n", encoding="utf-8")
        s = harness_v2.ingest(self.con, run_dir)
        self.assertEqual(s["mode"], "real")

    def test_mode_absent_defaults_unknown(self) -> None:
        run_dir = self.tmp_path / "T9-no-mode"
        run_dir.mkdir()
        (run_dir / "events.jsonl").write_text(
            json.dumps(_event_row(0, "run.start", {}, run_id="T9-no-mode")) + "\n",
            encoding="utf-8",
        )
        s = harness_v2.ingest(self.con, run_dir)
        self.assertEqual(s["mode"], "unknown")


class TestIngestAllDiscovery(_HarnessTestBase):

    def test_ingest_all_finds_every_run_dir(self) -> None:
        # Lay out three run-dirs under tmp/state/runs/.
        runs_root = self.tmp_path / "state" / "runs"
        runs_root.mkdir(parents=True)
        for name in ("Tx-one", "Ty-two", "Tz-three"):
            d = runs_root / name
            d.mkdir()
            (d / "events.jsonl").write_text(
                json.dumps(_event_row(0, "run.start", {}, run_id=name)) + "\n",
                encoding="utf-8",
            )
        # Decoy: a dir under runs/ that lacks events.jsonl must NOT be ingested.
        (runs_root / "decoy").mkdir()
        summaries = harness_v2.ingest_all(self.con, self.tmp_path / "state")
        ingested_ids = {s["run_id"] for s in summaries}
        self.assertEqual(ingested_ids, {"Tx-one", "Ty-two", "Tz-three"})


class TestXLSXExport(_HarnessTestBase):

    def test_xlsx_has_four_sheets_with_bold_headers(self) -> None:
        # v2 Phase 2 Step 2: export grew a fourth sheet
        # (validation_episodes) sourced from the new view.
        harness_v2.ingest(self.con, _CLEAN_DIR)
        harness_v2.ingest(self.con, _ESC_DIR)
        harness_v2.ingest(self.con, _RETRY_ESC_DIR)
        out = self.tmp_path / "out.xlsx"
        harness_v2.export_xlsx(self.con, out)
        self.assertTrue(out.is_file())
        wb = load_workbook(out)
        self.assertEqual(
            set(wb.sheetnames),
            {
                "operations",
                "per_run_summary",
                "per_task_comparison",
                "validation_episodes",
            },
        )
        # Header row 1 bold on every sheet, frozen panes at A2.
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            self.assertTrue(ws.cell(row=1, column=1).font.bold,
                            f"{sheet_name} header not bold")
            self.assertEqual(ws.freeze_panes, "A2",
                             f"{sheet_name} not frozen at A2")

    def test_xlsx_validation_episodes_sheet_header_order(self) -> None:
        """The validation_episodes sheet's header row must match
        `_VALIDATION_FAILURE_EPISODES_COLUMNS` 1:1 (column order is the
        contract the v2 Phase 2 exam reads against)."""
        harness_v2.ingest(self.con, _RETRY_ESC_DIR)
        out = self.tmp_path / "out.xlsx"
        harness_v2.export_xlsx(self.con, out)
        wb = load_workbook(out)
        ws = wb["validation_episodes"]
        actual_headers = tuple(
            ws.cell(row=1, column=ix).value
            for ix in range(1, len(harness_v2._VALIDATION_FAILURE_EPISODES_COLUMNS) + 1)
        )
        self.assertEqual(
            actual_headers,
            harness_v2._VALIDATION_FAILURE_EPISODES_COLUMNS,
        )
        # The single data row in the T2-fixture-only export carries the
        # expected episode shape.
        self.assertEqual(ws.cell(row=2, column=1).value, "T2-two-step-real")
        self.assertEqual(ws.cell(row=2, column=5).value, 1)  # n_validation_fails
        self.assertEqual(ws.cell(row=2, column=9).value, 1)  # extra_api_calls


class TestSelfCheck(unittest.TestCase):
    """The bundled self-check should pass without raising."""

    def test_self_check_passes(self) -> None:
        # Redirect ANVIL_ROOT so the production calibration.duckdb is
        # untouched by the test.
        with tempfile.TemporaryDirectory() as tmp:
            with mock.patch.dict(os.environ, {"ANVIL_ROOT": tmp}):
                rc = harness_v2.self_check()
        self.assertEqual(rc, 0)


# ---------------------------------------------------------------------------
# v2 Phase 3 Step 1 — spend_ledger (Candidate A) + cumulative_spend_by_task
# ---------------------------------------------------------------------------

class TestSpendLedger(_HarnessTestBase):
    """Candidate A: every ingest appends a ledger row; the prior
    non-superseded row for the (run_id, mode) is flipped to
    superseded=TRUE. History accumulates; the overwrite on
    events/run_metadata is unchanged."""

    def test_spend_ledger_initial_ingest_appends_one_row(self) -> None:
        harness_v2.ingest(self.con, _CLEAN_DIR)
        rows = self.con.execute(
            "SELECT run_id, mode, total_cost_usd, superseded FROM spend_ledger"
        ).fetchall()
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][0], "T1-doc-edit-mock")
        self.assertEqual(rows[0][1], "mock")
        self.assertGreater(rows[0][2], 0.0)  # non-zero cost snapshot
        self.assertFalse(rows[0][3])          # current (not superseded)

    def test_spend_ledger_second_ingest_supersedes_first(self) -> None:
        harness_v2.ingest(self.con, _CLEAN_DIR)
        harness_v2.ingest(self.con, _CLEAN_DIR)
        rows = self.con.execute(
            "SELECT ledger_id, total_cost_usd, superseded, ingest_ts "
            "FROM spend_ledger WHERE run_id='T1-doc-edit-mock' AND mode='mock' "
            "ORDER BY ledger_id"
        ).fetchall()
        self.assertEqual(len(rows), 2)
        # Lower ledger_id (the original) is superseded; higher is current.
        self.assertTrue(rows[0][2])
        self.assertFalse(rows[1][2])
        # ingest_ts is non-decreasing (≤ tolerates same-microsecond ties;
        # the superseded flag, not the ts, is the authoritative order).
        self.assertLessEqual(rows[0][3], rows[1][3])
        # cumulative_spend_by_task: n_ingests=2, cumulative=sum, latest=current.
        cum = self.con.execute(
            "SELECT n_ingests, cumulative_cost_usd, latest_cost_usd "
            "FROM cumulative_spend_by_task WHERE run_id='T1-doc-edit-mock'"
        ).fetchone()
        self.assertEqual(cum[0], 2)
        self.assertAlmostEqual(cum[1], rows[0][1] + rows[1][1], places=9)
        self.assertAlmostEqual(cum[2], rows[1][1], places=9)

    def test_spend_ledger_third_ingest_history_accumulates(self) -> None:
        for _ in range(3):
            harness_v2.ingest(self.con, _CLEAN_DIR)
        flags = [
            r[0] for r in self.con.execute(
                "SELECT superseded FROM spend_ledger "
                "WHERE run_id='T1-doc-edit-mock' ORDER BY ledger_id"
            ).fetchall()
        ]
        self.assertEqual(len(flags), 3)            # not capped at two
        self.assertEqual(flags.count(True), 2)     # two superseded
        self.assertEqual(flags.count(False), 1)    # one current
        self.assertFalse(flags[-1])                # current is the latest
        n = self.con.execute(
            "SELECT n_ingests FROM cumulative_spend_by_task "
            "WHERE run_id='T1-doc-edit-mock'"
        ).fetchone()[0]
        self.assertEqual(n, 3)

    def test_spend_ledger_transaction_atomicity(self) -> None:
        # First ingest commits one current row.
        harness_v2.ingest(self.con, _CLEAN_DIR)
        before = self.con.execute(
            "SELECT ledger_id, superseded FROM spend_ledger ORDER BY ledger_id"
        ).fetchall()
        self.assertEqual(len(before), 1)
        self.assertFalse(before[0][1])

        # Second ingest: inject a failure on the ledger INSERT (after the
        # superseded UPDATE). The whole transaction must roll back — the
        # UPDATE that flipped superseded=TRUE must not persist, and no row
        # is appended. Atomicity proof: the ledger is byte-identical to
        # `before`, never left with two non-superseded rows.
        #
        # DuckDBPyConnection.execute is a read-only C attribute, so it
        # can't be mock.patch'd. A thin proxy that delegates to the real
        # connection (same underlying transaction) but raises on the
        # ledger INSERT does the injection.
        class _FailOnLedgerInsert:
            def __init__(self, real):
                self._real = real

            def execute(self, sql, *a, **kw):
                if sql.strip().startswith("INSERT INTO spend_ledger"):
                    raise RuntimeError("injected failure mid-ledger-write")
                return self._real.execute(sql, *a, **kw)

            def __getattr__(self, name):
                return getattr(self._real, name)

        with self.assertRaises(RuntimeError):
            harness_v2.ingest(_FailOnLedgerInsert(self.con), _CLEAN_DIR)

        after = self.con.execute(
            "SELECT ledger_id, superseded FROM spend_ledger ORDER BY ledger_id"
        ).fetchall()
        self.assertEqual(after, before)  # rollback restored exact prior state

    def test_cumulative_spend_view_reconciliation(self) -> None:
        # Two ingests of T1-real (a re-run) + one of T2-real.
        harness_v2.ingest(self.con, _CLEAN_DIR_REAL)
        harness_v2.ingest(self.con, _CLEAN_DIR_REAL)
        harness_v2.ingest(self.con, _RETRY_ESC_DIR)
        # actual_real_spend = SUM over all real ledger rows (3 ingests).
        actual = self.con.execute(
            "SELECT SUM(total_cost_usd) FROM spend_ledger WHERE mode='real'"
        ).fetchone()[0]
        all_rows = self.con.execute(
            "SELECT total_cost_usd FROM spend_ledger WHERE mode='real'"
        ).fetchall()
        self.assertEqual(len(all_rows), 3)
        self.assertAlmostEqual(actual, sum(r[0] for r in all_rows), places=9)
        self.assertGreater(actual, 0.0)
        # recorded_latest = SUM per_run_summary real (latest only, 2 tasks).
        recorded = self.con.execute(
            "SELECT SUM(total_cost_usd) FROM per_run_summary WHERE mode='real'"
        ).fetchone()[0]
        # actual - recorded == the one superseded T1-real ingest.
        superseded_sum = self.con.execute(
            "SELECT COALESCE(SUM(total_cost_usd), 0.0) FROM spend_ledger "
            "WHERE mode='real' AND superseded = TRUE"
        ).fetchone()[0]
        self.assertAlmostEqual(actual - recorded, superseded_sum, places=9)


class TestHarnessCLI(_HarnessTestBase):
    """v2 Phase 3 Step 1 CLI surface: --db-path threading + the new
    cumulative-spend subcommand. All invocations pass --db-path to a tmp
    DB so the live v2-phase-2 default is never touched."""

    def test_harness_cli_db_path_flag(self) -> None:
        from contextlib import redirect_stdout
        import io
        # Regression: the bare-CLI default no longer points at the
        # pre-`mode` v1 DB (the binder-error root cause).
        self.assertIn("v2-phase-2", str(harness_v2.db_path()))
        self.assertNotIn("v2-phase-1", str(harness_v2.db_path()))
        # --db-path threads end-to-end into ingest + query.
        db = self.tmp_path / "cli.duckdb"
        with redirect_stdout(io.StringIO()):
            rc = harness_v2.main(["--db-path", str(db), "ingest", str(_CLEAN_DIR)])
        self.assertEqual(rc, 0)
        buf = io.StringIO()
        with redirect_stdout(buf):
            rc = harness_v2.main(
                ["--db-path", str(db), "per-run-summary", "T1-doc-edit-mock"]
            )
        self.assertEqual(rc, 0)
        self.assertIn("T1-doc-edit-mock", buf.getvalue())

    def test_operations_cli_positional_run_id(self) -> None:
        """v2 Phase 5 Step 1b: the `operations` subcommand takes a positional
        run_id (mirroring per-run-summary), so `operations T1-doc-edit-mock`
        works without a --run-id flag."""
        from contextlib import redirect_stdout
        import io
        db = self.tmp_path / "ops.duckdb"
        con = harness_v2.open_db(db)
        harness_v2.ingest(con, _CLEAN_DIR)
        con.close()
        buf = io.StringIO()
        with redirect_stdout(buf):
            rc = harness_v2.main(
                ["--db-path", str(db), "operations", "T1-doc-edit-mock"]
            )
        out = buf.getvalue()
        self.assertEqual(rc, 0)
        self.assertIn("T1-doc-edit-mock", out)
        # Bare `operations` (no run_id) prints all rows — the positional is
        # optional, so the run_id-filtered and unfiltered shapes both work.
        buf2 = io.StringIO()
        with redirect_stdout(buf2):
            rc2 = harness_v2.main(["--db-path", str(db), "operations"])
        self.assertEqual(rc2, 0)
        self.assertIn("T1-doc-edit-mock", buf2.getvalue())

    def test_harness_cli_cumulative_spend_subcommand(self) -> None:
        from contextlib import redirect_stdout
        import io
        db = self.tmp_path / "cli.duckdb"
        con = harness_v2.open_db(db)
        harness_v2.ingest(con, _CLEAN_DIR)
        harness_v2.ingest(con, _CLEAN_DIR)  # re-run → n_ingests=2
        con.close()
        buf = io.StringIO()
        with redirect_stdout(buf):
            rc = harness_v2.main(["--db-path", str(db), "cumulative-spend"])
        out = buf.getvalue()
        self.assertEqual(rc, 0)
        for col in harness_v2._CUMULATIVE_SPEND_COLUMNS:
            self.assertIn(col, out)
        self.assertIn("T1-doc-edit-mock", out)
        # --mode filter: a real filter excludes the mock-only fixture row.
        buf2 = io.StringIO()
        with redirect_stdout(buf2):
            rc2 = harness_v2.main(
                ["--db-path", str(db), "cumulative-spend", "--mode", "real"]
            )
        self.assertEqual(rc2, 0)
        self.assertNotIn("T1-doc-edit-mock", buf2.getvalue())


class TestV1Quarantine(unittest.TestCase):
    """v2 Phase 3 Step 1: the v1 DB is renamed to calibration-archived.duckdb
    so the (re-pointed) default CLI can't reach it. `state/` is gitignored,
    so a fresh checkout / CI carries neither file — skip there."""

    def test_v1_db_quarantine(self) -> None:
        anvil_root = Path(harness_v2.__file__).resolve().parent.parent
        original = anvil_root / "state" / "v2-phase-1" / "calibration.duckdb"
        archived = anvil_root / "state" / "v2-phase-1" / "calibration-archived.duckdb"
        if not original.exists() and not archived.exists():
            self.skipTest(
                "v1 DB absent in this environment (state/ is gitignored); "
                "quarantine verified manually at deployment"
            )
        self.assertFalse(
            original.exists(),
            "v1 calibration.duckdb still at original path — quarantine incomplete",
        )
        self.assertTrue(archived.exists(), "archived v1 DB missing")


class TestCoderCostFromReported(_HarnessTestBase):
    """v2 Phase 5 Step 1a: the operations view sources Coder cost from the
    CLI's *reported* total_cost_usd (the Coder runs a cheaper model than the
    Planner's Opus), NOT the Opus token-weighted formula — which would
    ~3x-over-cost it."""

    def test_operations_view_coder_cost_from_reported(self) -> None:
        run_dir = self.tmp_path / "T9-coder-cost"
        run_dir.mkdir()
        events = [
            _event_row(0, "run.start", {}, run_id="T9-coder-cost"),
            # Coder subprocess.end with v2 Phase 5 cost fields. The token
            # counts would, under the Opus formula, compute ~$0.147; the
            # reported figure is $0.04953 — the view must use the latter.
            _event_row(
                10, "coder.subprocess.end",
                {"exit_code": 0, "duration_ms": 1299, "stdout_chars": 30,
                 "stderr_chars": 0, "total_cost_usd": 0.04953225,
                 "input_tokens": 2152, "output_tokens": 4,
                 "cache_creation_input_tokens": 21000,
                 "cache_read_input_tokens": 10315},
                run_id="T9-coder-cost", step_idx=0,
            ),
            _event_row(20, "run.end", {}, run_id="T9-coder-cost"),
        ]
        (run_dir / "events.jsonl").write_text(
            "\n".join(json.dumps(e) for e in events) + "\n", encoding="utf-8",
        )
        (run_dir / "mode.txt").write_text("real\n", encoding="utf-8")
        harness_v2.ingest(self.con, run_dir)

        row = self.con.execute(
            "SELECT cost_usd, input_tokens, cache_read_tokens FROM operations "
            "WHERE run_id='T9-coder-cost' AND operation_kind='coder.subprocess.end'"
        ).fetchone()
        self.assertIsNotNone(row)
        # cost_usd = the reported figure, NOT the ~$0.147 formula value.
        # v3 Phase 1c Step 3.5: per-model Opus 4.7 rates $5/$25/$6.25/$0.50.
        self.assertAlmostEqual(row[0], 0.04953225, places=8)
        formula = (2152 * 5.0 + 4 * 25.0 + 21000 * 6.25 + 10315 * 0.50) / 1e6
        self.assertNotAlmostEqual(row[0], formula, places=3)  # ~0.147, rejected
        # token columns populated for observability (cost still reported).
        self.assertEqual(row[1], 2152)
        self.assertEqual(row[2], 10315)


class TestShadowDecisions(_HarnessTestBase):
    """v3 Phase 0 Step 2 (V3P0-3): shadow_decisions ingest + idempotency,
    and the champion_challenger_comparison view aggregation."""

    def _routing(self, stage, *, fallback=False):
        return {
            "route_candidate": "claude-opus-4-7",
            "route_actual": "claude-opus-4-7",
            "route_fallback_fired": fallback,
            "policy_version": "v3-phase-0-passive",
            "features_seen": {
                "observed_prompt_token_count": 500, "step_idx": 0,
                "stage": stage, "context_paths_count": 2,
            },
        }

    def _shadow_data(self, stage, *, agreement=True, actual="claude-opus-4-7"):
        return {
            "stage": stage,
            "shadow_route_candidate": "claude-opus-4-7",
            "shadow_decision_basis": {
                "observed_prompt_token_count": 500, "step_idx": 0,
                "stage": stage, "context_paths_count": 2,
            },
            "actual_route_taken": actual,
            "agreement": agreement,
        }

    def _ingest_shadow_run(self, run_id="T9-shadow", *, fallback=False,
                           b_agreement=True, b_actual="claude-opus-4-7"):
        run_dir = self.tmp_path / run_id
        run_dir.mkdir(exist_ok=True)
        evs = [
            _event_row(0, "run.start", {}, run_id=run_id),
            _event_row(10, "planner.stage_a.api_end",
                       {"model": "claude-opus-4-7", "input_tokens": 500,
                        "output_tokens": 50, "cache_creation_input_tokens": 0,
                        "cache_read_input_tokens": 0, "duration_ms": 900,
                        "ok": True, **self._routing("A")},
                       run_id=run_id, step_idx=0),
            _event_row(11, "shadow.decision", self._shadow_data("A"),
                       run_id=run_id, step_idx=0),
            _event_row(20, "planner.stage_b.api_end",
                       {"model": "claude-opus-4-7", "input_tokens": 500,
                        "output_tokens": 50, "cache_creation_input_tokens": 0,
                        "cache_read_input_tokens": 0, "duration_ms": 900,
                        "ok": True, **self._routing("B", fallback=fallback)},
                       run_id=run_id, step_idx=0),
            _event_row(21, "shadow.decision",
                       self._shadow_data("B", agreement=b_agreement,
                                         actual=b_actual),
                       run_id=run_id, step_idx=0),
            _event_row(30, "run.end", {"drops": 0}, run_id=run_id),
        ]
        (run_dir / "events.jsonl").write_text(
            "\n".join(json.dumps(e) for e in evs) + "\n", encoding="utf-8")
        (run_dir / "mode.txt").write_text("mock\n", encoding="utf-8")
        return harness_v2.ingest(self.con, run_dir)

    def test_shadow_decisions_ingest(self) -> None:
        self._ingest_shadow_run()
        rows = self.con.execute(
            "SELECT stage, shadow_route_candidate, actual_route_taken, "
            "       agreement, policy_version, "
            "       json_extract_string(shadow_decision_basis, '$.stage') "
            "FROM shadow_decisions WHERE run_id = 'T9-shadow' ORDER BY stage"
        ).fetchall()
        self.assertEqual(len(rows), 2)
        for r, stage in zip(rows, ("A", "B")):
            self.assertEqual(r[0], stage)
            self.assertEqual(r[1], "claude-opus-4-7")
            self.assertEqual(r[2], "claude-opus-4-7")
            self.assertIs(r[3], True)
            # policy_version column default applied at ingest (Phase 0).
            self.assertEqual(r[4], "v3-phase-0-passive")
            # shadow_decision_basis stored as JSON, reachable in-place.
            self.assertEqual(r[5], stage)

    def test_shadow_decisions_idempotent_reingest(self) -> None:
        # V2P2-2 (run_id, mode) composite-key invariant: re-ingest is
        # DELETE-then-INSERT, no duplicate rows.
        self._ingest_shadow_run()
        first = self.con.execute(
            "SELECT COUNT(*) FROM shadow_decisions WHERE run_id='T9-shadow'"
        ).fetchone()[0]
        self._ingest_shadow_run()
        second = self.con.execute(
            "SELECT COUNT(*) FROM shadow_decisions WHERE run_id='T9-shadow'"
        ).fetchone()[0]
        self.assertEqual(first, 2)
        self.assertEqual(second, 2)

    def test_champion_challenger_all_agree(self) -> None:
        self._ingest_shadow_run()
        rows = self.con.execute(
            "SELECT stage, policy_version, task_id, agreement_count, "
            "       disagreement_count, fallback_fired_count "
            "FROM champion_challenger_comparison "
            "WHERE task_id = 'T9' ORDER BY stage"
        ).fetchall()
        self.assertEqual(len(rows), 2)  # stage A, B
        for r, stage in zip(rows, ("A", "B")):
            self.assertEqual(r[0], stage)
            # policy_version populates from the operations join (the
            # synthetic api_end rows carry features_seen + policy_version).
            self.assertEqual(r[1], "v3-phase-0-passive")
            self.assertEqual(r[2], "T9")
            self.assertEqual(r[3], 1)   # agreement_count
            self.assertEqual(r[4], 0)   # disagreement_count
            self.assertEqual(r[5], 0)   # fallback_fired_count

    def test_champion_challenger_counts_disagreement_and_fallback(self) -> None:
        # Phase 0 never produces these, but the view's FILTER logic must
        # compute them correctly for Phase 1's first real shadow rule.
        self._ingest_shadow_run(fallback=True, b_agreement=False,
                                b_actual="claude-haiku-4-5")
        row_b = self.con.execute(
            "SELECT agreement_count, disagreement_count, fallback_fired_count "
            "FROM champion_challenger_comparison "
            "WHERE task_id = 'T9' AND stage = 'B'"
        ).fetchone()
        self.assertEqual(row_b[0], 0)   # no agreement (actual was haiku)
        self.assertEqual(row_b[1], 1)   # one disagreement
        self.assertEqual(row_b[2], 1)   # one fallback fired (on stage B api_end)

    def test_champion_challenger_joinable_to_per_task_comparison(self) -> None:
        # task_id is carried so the view joins to per_task_comparison.
        self._ingest_shadow_run()
        joined = self.con.execute(
            "SELECT cc.stage, cc.agreement_count "
            "FROM champion_challenger_comparison cc "
            "JOIN per_task_comparison ptc USING (task_id) "
            "WHERE cc.task_id = 'T9'"
        ).fetchall()
        # per_task_comparison has a T9 row (mock half present), so the
        # join returns the shadow aggregation rows.
        self.assertEqual(len(joined), 2)


class TestSilentMissEpisodes(_HarnessTestBase):
    """v3 Phase 0 Step 3 (V3P0-4): silent_miss_episodes is empty by
    construction in Phase 0 (silent_miss_count = 0) and materializes the
    moment a shadow_compare.end records a non-zero miss (Phase 1 shape)."""

    def _ingest_compare_run(self, run_id, silent_miss, *, hallucination=0,
                            jaccard=1.0):
        run_dir = self.tmp_path / run_id
        run_dir.mkdir(exist_ok=True)
        evs = [
            _event_row(0, "run.start", {}, run_id=run_id),
            _event_row(10, "planner.stage_a.api_end",
                       {"model": "claude-opus-4-7", "input_tokens": 500,
                        "output_tokens": 50, "duration_ms": 900, "ok": True},
                       run_id=run_id, step_idx=0),
            _event_row(11, "stage_a.shadow_compare.begin",
                       {"step_idx": 0, "routed_paths": [], "baseline_paths": []},
                       run_id=run_id, step_idx=0),
            _event_row(12, "stage_a.shadow_compare.end",
                       {"step_idx": 0, "silent_miss_count": silent_miss,
                        "hallucination_count": hallucination,
                        "jaccard_similarity": jaccard,
                        "baseline_only_paths": [], "routed_only_paths": []},
                       run_id=run_id, step_idx=0),
            _event_row(20, "run.end", {"drops": 0}, run_id=run_id),
        ]
        (run_dir / "events.jsonl").write_text(
            "\n".join(json.dumps(e) for e in evs) + "\n", encoding="utf-8")
        (run_dir / "mode.txt").write_text("mock\n", encoding="utf-8")
        return harness_v2.ingest(self.con, run_dir)

    def test_view_queryable_and_empty_in_phase_0(self) -> None:
        # Phase 0: silent_miss_count = 0 → the WHERE > 0 filter yields no
        # rows, but the view is queryable.
        self._ingest_compare_run("T9-nomiss", silent_miss=0)
        rows = self.con.execute(
            "SELECT * FROM silent_miss_episodes"
        ).fetchall()
        self.assertEqual(len(rows), 0)

    def test_view_materializes_on_nonzero_miss(self) -> None:
        # Phase 1 shape: a non-zero silent_miss surfaces one episode row,
        # joined to run_metadata for task_id.
        self._ingest_compare_run("T9-miss", silent_miss=3, hallucination=1,
                                 jaccard=0.5)
        rows = self.con.execute(
            "SELECT run_id, task_id, mode, step_idx, silent_miss_count, "
            "       hallucination_count, jaccard_similarity "
            "FROM silent_miss_episodes WHERE run_id = 'T9-miss'"
        ).fetchall()
        self.assertEqual(len(rows), 1)
        r = rows[0]
        self.assertEqual(r[0], "T9-miss")
        self.assertEqual(r[1], "T9")          # task_id via run_metadata join
        self.assertEqual(r[2], "mock")
        self.assertEqual(r[3], 0)
        self.assertEqual(r[4], 3)             # silent_miss_count
        self.assertEqual(r[5], 1)             # hallucination_count
        self.assertEqual(r[6], 0.5)           # jaccard_similarity

    def test_columns_match_declared_tuple(self) -> None:
        self._ingest_compare_run("T9-miss2", silent_miss=1)
        row = self.con.execute(
            "SELECT * FROM silent_miss_episodes WHERE run_id='T9-miss2'"
        ).fetchone()
        self.assertEqual(len(row), len(harness_v2._SILENT_MISS_EPISODES_COLUMNS))


class TestCacheDiagnosticsView(_HarnessTestBase):
    """v3 Phase 0 Step 4 (V3P0-6): cache_diagnostics exposes the three
    cache-family lines per Planner stage, with vault_index_hit null on
    Stage B/C and a summed block-size convenience column."""

    def _ingest_cache_run(self, run_id="T9-cache"):
        run_dir = self.tmp_path / run_id
        run_dir.mkdir(exist_ok=True)
        a_blocks = {"brief": 820, "state": 130, "vault_files": 45, "prior_step": 0}
        b_blocks = {"brief": 820, "state": 160, "vault_files": 9200, "prior_step": 60}
        evs = [
            _event_row(0, "run.start", {}, run_id=run_id),
            _event_row(10, "planner.stage_a.api_end",
                       {"model": "claude-opus-4-7", "input_tokens": 1040,
                        "vault_index_hit": False,
                        "candidate_user_block_sizes": a_blocks,
                        "seconds_since_cache_creation": None, "ok": True},
                       run_id=run_id, step_idx=0),
            _event_row(20, "planner.stage_b.api_end",
                       {"model": "claude-opus-4-7", "input_tokens": 13719,
                        "vault_index_hit": None,
                        "candidate_user_block_sizes": b_blocks,
                        "seconds_since_cache_creation": 2.5, "ok": True},
                       run_id=run_id, step_idx=0),
            _event_row(30, "run.end", {"drops": 0}, run_id=run_id),
        ]
        (run_dir / "events.jsonl").write_text(
            "\n".join(json.dumps(e) for e in evs) + "\n", encoding="utf-8")
        (run_dir / "mode.txt").write_text("mock\n", encoding="utf-8")
        return harness_v2.ingest(self.con, run_dir)

    def test_stage_a_row(self) -> None:
        self._ingest_cache_run()
        row = self.con.execute(
            "SELECT task_id, mode, vault_index_hit, "
            "       json_extract_string(candidate_user_block_sizes, '$.vault_files'), "
            "       seconds_since_cache_creation, candidate_block_sizes_sum, input_tokens "
            "FROM cache_diagnostics WHERE run_id='T9-cache' AND stage='A'"
        ).fetchone()
        self.assertEqual(row[0], "T9")          # task_id via run_metadata join
        self.assertEqual(row[1], "mock")
        self.assertIs(row[2], False)            # vault_index_hit (Stage A)
        self.assertEqual(row[3], "45")          # block sizes queryable as JSON
        self.assertIsNone(row[4])               # seconds null (creation call)
        self.assertEqual(row[5], 820 + 130 + 45 + 0)   # sum
        self.assertEqual(row[6], 1040)

    def test_stage_b_vault_index_hit_null(self) -> None:
        self._ingest_cache_run()
        row = self.con.execute(
            "SELECT vault_index_hit, seconds_since_cache_creation, "
            "       candidate_block_sizes_sum "
            "FROM cache_diagnostics WHERE run_id='T9-cache' AND stage='B'"
        ).fetchone()
        self.assertIsNone(row[0])               # null on Stage B (Q(c))
        self.assertEqual(row[1], 2.5)           # read within window
        self.assertEqual(row[2], 820 + 160 + 9200 + 60)

    def test_uncached_user_prompt_equiv_columns(self) -> None:
        # v3 Phase 1c Step 1 (Step1C-F1): the view exposes model,
        # cache_read/creation, and the cache-invariant uncached_user_prompt_equiv.
        run_id = "T10-equiv"
        run_dir = self.tmp_path / run_id
        run_dir.mkdir(exist_ok=True)
        blocks = {"brief": 400, "state": 300, "vault_files": 0, "prior_step": 0}
        evs = [
            _event_row(0, "run.start", {}, run_id=run_id),
            # Cache-READ Opus call: the 3479-token system prompt is in
            # cache_read; input_tokens holds only the user prompt.
            _event_row(10, "planner.stage_a.api_end",
                       {"model": "claude-opus-4-7", "input_tokens": 1500,
                        "cache_read_input_tokens": 3479,
                        "cache_creation_input_tokens": 0,
                        "vault_index_hit": False,
                        "candidate_user_block_sizes": blocks,
                        "seconds_since_cache_creation": 1.0, "ok": True},
                       run_id=run_id, step_idx=0),
            _event_row(30, "run.end", {"drops": 0}, run_id=run_id),
        ]
        (run_dir / "events.jsonl").write_text(
            "\n".join(json.dumps(e) for e in evs) + "\n", encoding="utf-8")
        (run_dir / "mode.txt").write_text("real\n", encoding="utf-8")
        harness_v2.ingest(self.con, run_dir)
        row = self.con.execute(
            "SELECT model, cache_read_input_tokens, cache_creation_input_tokens, "
            "       uncached_user_prompt_equiv "
            "FROM cache_diagnostics WHERE run_id='T10-equiv' AND stage='A'"
        ).fetchone()
        self.assertEqual(row[0], "claude-opus-4-7")
        self.assertEqual(row[1], 3479)
        self.assertEqual(row[2], 0)
        # equiv = input_tokens + cache_read + cache_creation - 3479 = 1500
        self.assertEqual(row[3], 1500)

    def test_uncached_user_prompt_equiv_per_model_haiku(self) -> None:
        # v3 Phase 2a Step 1 (V3P2A-1, Q-A5): the per-model CASE subtracts the
        # HAIKU system constant (2590), not the Opus 3479. A Haiku Stage A row
        # (cr=cc=0 — Haiku doesn't cache, Step3.5C-F3): equiv = 3550 - 2590 =
        # 960. (At the old Opus-only 3479 it would be a nonsense 71 — the
        # Step1C-F2 footgun this step fixes.)
        run_id = "T12-haiku-equiv"
        run_dir = self.tmp_path / run_id
        run_dir.mkdir(exist_ok=True)
        blocks = {"brief": 400, "state": 88, "vault_files": 0, "prior_step": 0}
        evs = [
            _event_row(0, "run.start", {}, run_id=run_id),
            _event_row(10, "planner.stage_a.api_end",
                       {"model": "claude-haiku-4-5-20251001",
                        "input_tokens": 3550,
                        "cache_read_input_tokens": 0,
                        "cache_creation_input_tokens": 0,
                        "vault_index_hit": False,
                        "candidate_user_block_sizes": blocks,
                        "seconds_since_cache_creation": None, "ok": True},
                       run_id=run_id, step_idx=0),
            _event_row(30, "run.end", {"drops": 0}, run_id=run_id),
        ]
        (run_dir / "events.jsonl").write_text(
            "\n".join(json.dumps(e) for e in evs) + "\n", encoding="utf-8")
        (run_dir / "mode.txt").write_text("real\n", encoding="utf-8")
        harness_v2.ingest(self.con, run_dir)
        row = self.con.execute(
            "SELECT model, uncached_user_prompt_equiv "
            "FROM cache_diagnostics WHERE run_id='T12-haiku-equiv' AND stage='A'"
        ).fetchone()
        self.assertEqual(row[0], "claude-haiku-4-5-20251001")
        self.assertEqual(row[1], 3550 - 2590)   # 960, per-model (NOT 3550-3479)

    def test_uncached_user_prompt_equiv_unknown_model_opus_fallback(self) -> None:
        # Unknown model → the CASE ELSE branch (Opus 3479, conservative).
        run_id = "T13-unknown-equiv"
        run_dir = self.tmp_path / run_id
        run_dir.mkdir(exist_ok=True)
        blocks = {"brief": 400, "state": 0, "vault_files": 0, "prior_step": 0}
        evs = [
            _event_row(0, "run.start", {}, run_id=run_id),
            _event_row(10, "planner.stage_a.api_end",
                       {"model": "claude-future-9", "input_tokens": 4000,
                        "cache_read_input_tokens": 0,
                        "cache_creation_input_tokens": 0,
                        "vault_index_hit": False,
                        "candidate_user_block_sizes": blocks,
                        "seconds_since_cache_creation": None, "ok": True},
                       run_id=run_id, step_idx=0),
            _event_row(30, "run.end", {"drops": 0}, run_id=run_id),
        ]
        (run_dir / "events.jsonl").write_text(
            "\n".join(json.dumps(e) for e in evs) + "\n", encoding="utf-8")
        (run_dir / "mode.txt").write_text("real\n", encoding="utf-8")
        harness_v2.ingest(self.con, run_dir)
        equiv = self.con.execute(
            "SELECT uncached_user_prompt_equiv FROM cache_diagnostics "
            "WHERE run_id='T13-unknown-equiv' AND stage='A'"
        ).fetchone()[0]
        self.assertEqual(equiv, 4000 - 3479)   # Opus fallback (ELSE branch)

    def test_brief_cache_read_pattern_surfaces(self) -> None:
        # v3 Phase 1c Step 2 (Q10.9): the brief cache pattern — cache_creation
        # on step 0 (read=0), cache_read>0 on step 1 (the [system+brief] prefix
        # hit) — surfaces per-step in cache_diagnostics.
        run_id = "T11-briefcache"
        run_dir = self.tmp_path / run_id
        run_dir.mkdir(exist_ok=True)
        blocks = {"brief": 700, "state": 200, "vault_files": 0, "prior_step": 0}
        evs = [
            _event_row(0, "run.start", {}, run_id=run_id),
            _event_row(10, "planner.stage_a.api_end",
                       {"model": "claude-opus-4-7", "input_tokens": 900,
                        "cache_read_input_tokens": 0,
                        "cache_creation_input_tokens": 4180,
                        "vault_index_hit": False,
                        "candidate_user_block_sizes": blocks,
                        "seconds_since_cache_creation": None, "ok": True},
                       run_id=run_id, step_idx=0),
            _event_row(20, "planner.stage_a.api_end",
                       {"model": "claude-opus-4-7", "input_tokens": 200,
                        "cache_read_input_tokens": 4180,
                        "cache_creation_input_tokens": 0,
                        "vault_index_hit": True,
                        "candidate_user_block_sizes": blocks,
                        "seconds_since_cache_creation": 2.0, "ok": True},
                       run_id=run_id, step_idx=1),
            _event_row(30, "run.end", {"drops": 0}, run_id=run_id),
        ]
        (run_dir / "events.jsonl").write_text(
            "\n".join(json.dumps(e) for e in evs) + "\n", encoding="utf-8")
        (run_dir / "mode.txt").write_text("real\n", encoding="utf-8")
        harness_v2.ingest(self.con, run_dir)
        rows = self.con.execute(
            "SELECT step_idx, cache_read_input_tokens, cache_creation_input_tokens, "
            "       uncached_user_prompt_equiv "
            "FROM cache_diagnostics WHERE run_id='T11-briefcache' AND stage='A' "
            "ORDER BY step_idx"
        ).fetchall()
        self.assertEqual(len(rows), 2)
        self.assertEqual((rows[0][1], rows[0][2]), (0, 4180))      # step 0: creation
        self.assertEqual((rows[1][1], rows[1][2]), (4180, 0))      # step 1: read
        self.assertEqual(rows[1][3], 200 + 4180 + 0 - 3479)        # equiv computes

    def test_columns_match_declared_tuple(self) -> None:
        self._ingest_cache_run()
        row = self.con.execute(
            "SELECT * FROM cache_diagnostics WHERE run_id='T9-cache' AND stage='A'"
        ).fetchone()
        self.assertEqual(len(row), len(harness_v2._CACHE_DIAGNOSTICS_COLUMNS))

    def test_view_per_stage_and_mode(self) -> None:
        self._ingest_cache_run()
        rows = self.con.execute(
            "SELECT stage FROM cache_diagnostics WHERE run_id='T9-cache' "
            "ORDER BY stage"
        ).fetchall()
        self.assertEqual([r[0] for r in rows], ["A", "B"])


class TestPerModelSystemPromptTokensSingleSource(unittest.TestCase):
    """v3 Phase 2a Step 1 (V3P2A-1): the cache_diagnostics per-model
    system-prompt CASE generates from anvil.events.PLANNER_SYSTEM_PROMPT_
    TOKENS_BY_MODEL — the single source of truth. harness_v2 IMPORTS the
    mapping (not a re-hardcoded literal), so the two cannot drift.

    Step-0 audit nuance: exam_harness does NOT consume the token constants
    (only MODEL_RATES), so the genuine anti-drift mirror is events ↔
    harness_v2, not the brief-assumed harness_v2 ↔ exam_harness."""

    def test_harness_imports_events_mapping_no_drift(self) -> None:
        # Same object — single source of truth, structurally drift-proof.
        self.assertIs(
            harness_v2.PLANNER_SYSTEM_PROMPT_TOKENS_BY_MODEL,
            events.PLANNER_SYSTEM_PROMPT_TOKENS_BY_MODEL)
        self.assertEqual(
            events.PLANNER_SYSTEM_PROMPT_TOKENS_BY_MODEL,
            {"claude-opus-4-7": 3479, "claude-haiku-4-5-20251001": 2590})

    def test_case_sql_covers_every_model_with_opus_default(self) -> None:
        sql = harness_v2._system_prompt_tokens_case_sql()
        for model, tokens in events.PLANNER_SYSTEM_PROMPT_TOKENS_BY_MODEL.items():
            self.assertIn(f"WHEN '{model}' THEN {tokens}", sql)
        opus = events.PLANNER_SYSTEM_PROMPT_TOKENS_BY_MODEL["claude-opus-4-7"]
        self.assertIn(f"ELSE {opus} END", sql)


class TestStageASelectionsView(_HarnessTestBase):
    """v3 Phase 2a Step 2 (V3P2A-2): stage_a_selections exposes the recorded
    selection list + raw response. selected_paths/raw_response_text/truncated
    live in the events.data JSON (no per-field ingest column — same store as
    candidate_user_block_sizes); the view projects them. The list round-trips
    as a JSON array: content AND order preserved."""

    def _ingest_parsed(self, run_id, *, selected, raw, truncated, dropped=0):
        run_dir = self.tmp_path / run_id
        run_dir.mkdir(exist_ok=True)
        evs = [
            _event_row(0, "run.start", {}, run_id=run_id),
            _event_row(10, "planner.stage_a.parsed",
                       {"step_idx": 0, "paths_returned": len(selected),
                        "paths_dropped_as_hallucinated": dropped,
                        "selected_paths": selected,
                        "raw_response_text": raw, "truncated": truncated},
                       run_id=run_id, step_idx=0),
            _event_row(30, "run.end", {"drops": 0}, run_id=run_id),
        ]
        (run_dir / "events.jsonl").write_text(
            "\n".join(json.dumps(e) for e in evs) + "\n", encoding="utf-8")
        (run_dir / "mode.txt").write_text("real\n", encoding="utf-8")
        harness_v2.ingest(self.con, run_dir)

    def test_selected_paths_roundtrip_content_and_order(self) -> None:
        sel = ["docs/c.md", "docs/a.md", "docs/b.md"]  # deliberately not sorted
        self._ingest_parsed("T20-sel", selected=sel,
                            raw="docs/c.md\ndocs/a.md\ndocs/b.md", truncated=False)
        row = self.con.execute(
            "SELECT selected_paths, paths_returned FROM stage_a_selections "
            "WHERE run_id='T20-sel'").fetchone()
        self.assertEqual(json.loads(row[0]), sel)   # content AND order preserved
        self.assertEqual(row[1], 3)

    def test_empty_selection_roundtrips_as_empty_list(self) -> None:
        self._ingest_parsed("T21-empty", selected=[], raw="", truncated=False)
        row = self.con.execute(
            "SELECT selected_paths, paths_returned FROM stage_a_selections "
            "WHERE run_id='T21-empty'").fetchone()
        self.assertEqual(json.loads(row[0]), [])     # [] not null
        self.assertEqual(row[1], 0)

    def test_raw_response_and_truncated_false_roundtrip(self) -> None:
        self._ingest_parsed("T22-raw", selected=[], raw="the model said this",
                            truncated=False)
        row = self.con.execute(
            "SELECT raw_response_text, truncated FROM stage_a_selections "
            "WHERE run_id='T22-raw'").fetchone()
        self.assertEqual(row[0], "the model said this")
        self.assertIs(row[1], False)

    def test_truncated_flag_true_roundtrips(self) -> None:
        self._ingest_parsed("T23-trunc", selected=[], raw="z" * 16384,
                            truncated=True)
        row = self.con.execute(
            "SELECT truncated, length(raw_response_text) FROM stage_a_selections "
            "WHERE run_id='T23-trunc'").fetchone()
        self.assertIs(row[0], True)
        self.assertEqual(row[1], 16384)

    def test_paths_returned_equals_selected_len_at_view(self) -> None:
        self._ingest_parsed("T24-inv", selected=["a.md", "b.md"],
                            raw="a.md\nb.md", truncated=False)
        row = self.con.execute(
            "SELECT paths_returned, json_array_length(selected_paths) "
            "FROM stage_a_selections WHERE run_id='T24-inv'").fetchone()
        self.assertEqual(row[0], row[1])   # back-compat invariant at the view

    def test_columns_match_declared_tuple(self) -> None:
        self._ingest_parsed("T25-cols", selected=["a.md"], raw="a.md",
                            truncated=False)
        row = self.con.execute(
            "SELECT * FROM stage_a_selections WHERE run_id='T25-cols'").fetchone()
        self.assertEqual(len(row), len(harness_v2._STAGE_A_SELECTIONS_COLUMNS))


class TestPerStageRouteActual(_HarnessTestBase):
    """v3 Phase 1a Step 1: the operations view surfaces route_actual
    distinctly per stage now that a Planner can route Stage C to a
    different model than Stage A/B. The harness is pure projection (no
    functional change this step); this guards that the per-stage
    attribution reaches the operations.route_actual column intact."""

    def _routing(self, route):
        return {
            "route_candidate": route,
            "route_actual": route,
            "route_fallback_fired": False,
            "policy_version": "v3-phase-0-passive",
        }

    def _api_end(self, kind, route, *, stage, step_idx):
        data = {
            "model": route, "input_tokens": 1000, "output_tokens": 50,
            "cache_creation_input_tokens": 0, "cache_read_input_tokens": 0,
            "duration_ms": 1500, "ok": True,
            "features_seen": {"stage": stage},
        }
        data.update(self._routing(route))
        return _event_row(10, kind, data, run_id="T9-perstage", step_idx=step_idx)

    def test_operations_route_actual_differs_per_stage(self) -> None:
        run_dir = self.tmp_path / "T9-perstage"
        run_dir.mkdir()
        rows = [
            _event_row(0, "run.start", {}, run_id="T9-perstage"),
            self._api_end("planner.stage_a.api_end", "claude-opus-4-7",
                          stage="A", step_idx=0),
            self._api_end("planner.stage_b.api_end", "claude-opus-4-7",
                          stage="B", step_idx=0),
            self._api_end("planner.stage_c.api_end", "claude-sonnet-4-6",
                          stage="C", step_idx=None),
            _event_row(20, "run.end", {"drops": 0}, run_id="T9-perstage"),
        ]
        (run_dir / "events.jsonl").write_text(
            "\n".join(json.dumps(e) for e in rows) + "\n", encoding="utf-8",
        )
        (run_dir / "mode.txt").write_text("mock\n", encoding="utf-8")
        harness_v2.ingest(self.con, run_dir)

        result = self.con.execute(
            "SELECT operation_kind, route_actual FROM operations "
            "WHERE run_id = 'T9-perstage' ORDER BY operation_kind"
        ).fetchall()
        by_kind = {k: ra for k, ra in result}
        self.assertEqual(by_kind["planner.stage_a.api_end"], "claude-opus-4-7")
        self.assertEqual(by_kind["planner.stage_b.api_end"], "claude-opus-4-7")
        self.assertEqual(by_kind["planner.stage_c.api_end"], "claude-sonnet-4-6")


class TestStep3PolicyVersion(_HarnessTestBase):
    """v3 Phase 1a Step 3: the shadow_decisions ingest reads policy_version from
    the event (criterion 4), and champion_challenger_comparison segregates the
    two policy generations into separate rows (criterion 5)."""

    def test_shadow_ingest_reads_policy_version(self) -> None:
        run_dir = self.tmp_path / "T9-pv"
        run_dir.mkdir()
        rows = [
            _event_row(0, "run.start", {}, run_id="T9-pv"),
            _event_row(
                10, "shadow.decision",
                {"stage": "A", "shadow_route_candidate": "claude-opus-4-7",
                 "shadow_decision_basis": {"stage": "A"},
                 "actual_route_taken": "claude-opus-4-7", "agreement": True,
                 "policy_version": "v3-phase-1a-placeholder"},
                run_id="T9-pv", step_idx=0),
            _event_row(
                20, "shadow.decision",
                {"stage": "B", "shadow_route_candidate": "claude-opus-4-7",
                 "shadow_decision_basis": {"stage": "B"},
                 "actual_route_taken": "claude-opus-4-7", "agreement": True},
                run_id="T9-pv", step_idx=0),  # no policy_version → column default
            _event_row(30, "run.end", {"drops": 0}, run_id="T9-pv"),
        ]
        (run_dir / "events.jsonl").write_text(
            "\n".join(json.dumps(e) for e in rows) + "\n", encoding="utf-8")
        (run_dir / "mode.txt").write_text("mock\n", encoding="utf-8")
        harness_v2.ingest(self.con, run_dir)

        got = dict(self.con.execute(
            "SELECT stage, policy_version FROM shadow_decisions "
            "WHERE run_id = 'T9-pv' ORDER BY stage"
        ).fetchall())
        self.assertEqual(got["A"], "v3-phase-1a-placeholder")  # read from event
        self.assertEqual(got["B"], "v3-phase-0-passive")       # column default

    def _make_policy_run(self, run_id: str, policy_version: str) -> None:
        run_dir = self.tmp_path / run_id
        run_dir.mkdir()
        api_data = {
            "model": "claude-opus-4-7", "input_tokens": 1000, "output_tokens": 50,
            "cache_creation_input_tokens": 0, "cache_read_input_tokens": 0,
            "duration_ms": 1500, "ok": True,
            "route_candidate": "claude-opus-4-7", "route_actual": "claude-opus-4-7",
            "route_fallback_fired": False, "policy_version": policy_version,
            "features_seen": {"stage": "A", "step_idx": 0,
                              "observed_prompt_token_count": 1000,
                              "context_paths_count": 0},
        }
        shadow_data = {
            "stage": "A", "shadow_route_candidate": "claude-opus-4-7",
            "shadow_decision_basis": {"stage": "A"},
            "actual_route_taken": "claude-opus-4-7", "agreement": True,
            "policy_version": policy_version,
        }
        rows = [
            _event_row(0, "run.start", {}, run_id=run_id),
            _event_row(10, "planner.stage_a.api_end", api_data,
                       run_id=run_id, step_idx=0),
            _event_row(11, "shadow.decision", shadow_data,
                       run_id=run_id, step_idx=0),
            _event_row(20, "run.end", {"drops": 0}, run_id=run_id),
        ]
        (run_dir / "events.jsonl").write_text(
            "\n".join(json.dumps(e) for e in rows) + "\n", encoding="utf-8")
        (run_dir / "mode.txt").write_text("mock\n", encoding="utf-8")
        harness_v2.ingest(self.con, run_dir)

    def test_champion_challenger_returns_both_policy_versions(self) -> None:
        # Distinct task_ids (T8 vs T9) so the two policy generations aggregate
        # as SEPARATE rows rather than merging into one (stage, task_id) group.
        self._make_policy_run("T8-passive", "v3-phase-0-passive")
        self._make_policy_run("T9-placeholder", "v3-phase-1a-placeholder")

        rows = self.con.execute(
            "SELECT policy_version, agreement_count, disagreement_count "
            "FROM champion_challenger_comparison "
            "WHERE policy_version IS NOT NULL"
        ).fetchall()
        by_pv = {pv: (ac, dc) for pv, ac, dc in rows}
        self.assertIn("v3-phase-0-passive", by_pv)
        self.assertIn("v3-phase-1a-placeholder", by_pv)
        # 100% agreement on each: agreement_count > 0, disagreement_count == 0.
        for pv in ("v3-phase-0-passive", "v3-phase-1a-placeholder"):
            ac, dc = by_pv[pv]
            self.assertGreater(ac, 0)
            self.assertEqual(dc, 0)

    def _make_cc_run(self, run_id, policy_version, candidate, actual) -> None:
        # v3 Phase 1b Step 2: a run whose Stage A shadow.decision can DISAGREE
        # (candidate != actual) — the placeholder helper always agrees.
        run_dir = self.tmp_path / run_id
        run_dir.mkdir()
        agreement = candidate == actual
        api_data = {
            "model": actual, "input_tokens": 1000, "output_tokens": 50,
            "cache_creation_input_tokens": 0, "cache_read_input_tokens": 0,
            "duration_ms": 1500, "ok": True,
            "route_candidate": candidate, "route_actual": actual,
            "route_fallback_fired": False, "policy_version": policy_version,
            "features_seen": {"stage": "A", "step_idx": 0,
                              "observed_prompt_token_count": 1000,
                              "context_paths_count": 0},
        }
        shadow_data = {
            "stage": "A", "shadow_route_candidate": candidate,
            "shadow_decision_basis": {"stage": "A"},
            "actual_route_taken": actual, "agreement": agreement,
            "policy_version": policy_version,
        }
        rows = [
            _event_row(0, "run.start", {}, run_id=run_id),
            _event_row(10, "planner.stage_a.api_end", api_data,
                       run_id=run_id, step_idx=0),
            _event_row(11, "shadow.decision", shadow_data,
                       run_id=run_id, step_idx=0),
            _event_row(20, "run.end", {"drops": 0}, run_id=run_id),
        ]
        (run_dir / "events.jsonl").write_text(
            "\n".join(json.dumps(e) for e in rows) + "\n", encoding="utf-8")
        (run_dir / "mode.txt").write_text("mock\n", encoding="utf-8")
        harness_v2.ingest(self.con, run_dir)

    def test_shadow_ingest_reads_stage_a_shadow_policy_version(self) -> None:
        self._make_cc_run("T7-sh", "v3-phase-1b-stage-a-shadow",
                          "claude-haiku-4-5-20251001", "claude-opus-4-7")
        got = self.con.execute(
            "SELECT DISTINCT policy_version FROM shadow_decisions "
            "WHERE run_id = 'T7-sh'"
        ).fetchall()
        self.assertEqual(got, [("v3-phase-1b-stage-a-shadow",)])

    def test_champion_challenger_disagreement_on_shadow_agreement_on_placeholder(
        self,
    ) -> None:
        # Criterion 4 contract proof: in ONE DuckDB, the placeholder rows show
        # agreement and the shadow rows show disagreement (Haiku candidate ≠ Opus
        # actual). Distinct task_ids so they aggregate as separate rows.
        self._make_cc_run("T8-ph", "v3-phase-1a-placeholder",
                          "claude-opus-4-7", "claude-opus-4-7")
        self._make_cc_run("T9-sh", "v3-phase-1b-stage-a-shadow",
                          "claude-haiku-4-5-20251001", "claude-opus-4-7")
        rows = self.con.execute(
            "SELECT policy_version, SUM(agreement_count), SUM(disagreement_count) "
            "FROM champion_challenger_comparison WHERE policy_version IS NOT NULL "
            "GROUP BY policy_version"
        ).fetchall()
        by_pv = {pv: (ac, dc) for pv, ac, dc in rows}
        self.assertEqual(by_pv["v3-phase-1a-placeholder"], (1, 0))    # agreement
        self.assertEqual(by_pv["v3-phase-1b-stage-a-shadow"], (0, 1))  # disagreement

    def _ingest_canary_baseline(self, run_id) -> None:
        run_dir = self.tmp_path / run_id
        run_dir.mkdir()
        baseline = {
            "step_idx": 0, "model": "claude-opus-4-7",
            "input_tokens": 1000, "output_tokens": 50,
            "cache_creation_input_tokens": 0, "cache_read_input_tokens": 0,
            "duration_ms": 1500, "ok": True,
        }
        rows = [
            _event_row(0, "run.start", {}, run_id=run_id),
            _event_row(10, "planner.stage_a.canary_baseline.api_end", baseline,
                       run_id=run_id, step_idx=0),
            _event_row(20, "run.end", {"drops": 0}, run_id=run_id),
        ]
        (run_dir / "events.jsonl").write_text(
            "\n".join(json.dumps(e) for e in rows) + "\n", encoding="utf-8")
        (run_dir / "mode.txt").write_text("real\n", encoding="utf-8")
        harness_v2.ingest(self.con, run_dir)

    def test_canary_baseline_appears_in_operations_with_cost(self) -> None:
        # v3 Phase 1b Step 3: the parallel-Opus baseline is a counted operation,
        # its cost computed by the token formula (like any non-coder api_end).
        self._ingest_canary_baseline("T1-cb-ops")
        row = self.con.execute(
            "SELECT operation_kind, cost_usd FROM operations "
            "WHERE run_id = 'T1-cb-ops' "
            "AND operation_kind = 'planner.stage_a.canary_baseline.api_end'"
        ).fetchone()
        self.assertIsNotNone(row)
        self.assertGreater(row[1], 0.0)  # 1000*15 + 50*75 over 1e6 = $0.01875

    def test_canary_baseline_cost_in_per_run_summary(self) -> None:
        # Spend reconciliation (criterion 5 carry-forward): the baseline's cost
        # is ledgered into the per-run total.
        self._ingest_canary_baseline("T1-cb-sum")
        total = self.con.execute(
            "SELECT total_cost_usd FROM per_run_summary WHERE run_id = 'T1-cb-sum'"
        ).fetchone()[0]
        self.assertGreater(total, 0.0)


class TestPerModelCostRates(_HarnessTestBase):
    """v3 Phase 1c Step 3.5 (Step3.5C-F1): the operations-view cost formula
    uses per-model rates (MODEL_RATES) — Opus 4.7 $5/$25/$6.25/$0.50 and Haiku
    4.5 $1/$5/$1.25/$0.10 — instead of a single (stale Opus-4.1) table."""

    def _ingest_planner_event(self, run_id, model, inp, outp, cc, cr,
                              kind="planner.stage_a.api_end", mode="real"):
        run_dir = self.tmp_path / run_id
        run_dir.mkdir(exist_ok=True)
        evs = [
            _event_row(0, "run.start", {}, run_id=run_id),
            _event_row(10, kind,
                       {"model": model, "input_tokens": inp, "output_tokens": outp,
                        "cache_creation_input_tokens": cc,
                        "cache_read_input_tokens": cr, "ok": True},
                       run_id=run_id, step_idx=0),
            _event_row(20, "run.end", {}, run_id=run_id),
        ]
        (run_dir / "events.jsonl").write_text(
            "\n".join(json.dumps(e) for e in evs) + "\n", encoding="utf-8")
        (run_dir / "mode.txt").write_text(f"{mode}\n", encoding="utf-8")
        return harness_v2.ingest(self.con, run_dir)

    def _cost(self, run_id):
        return self.con.execute(
            "SELECT cost_usd FROM operations WHERE run_id = ? "
            "AND operation_kind = 'planner.stage_a.api_end'", [run_id]
        ).fetchone()[0]

    def test_model_rates_table_entries(self) -> None:
        # Q8.1: MODEL_RATES carries the two v3 models at the real rates.
        self.assertEqual(harness_v2.MODEL_RATES["claude-opus-4-7"],
                         {"input": 5.0, "output": 25.0,
                          "cache_create": 6.25, "cache_read": 0.50})
        self.assertEqual(harness_v2.MODEL_RATES["claude-haiku-4-5-20251001"],
                         {"input": 1.0, "output": 5.0,
                          "cache_create": 1.25, "cache_read": 0.10})
        self.assertIs(harness_v2.DEFAULT_MODEL_RATES,
                      harness_v2.MODEL_RATES["claude-opus-4-7"])

    def test_opus_row_cost(self) -> None:
        # Q8.2: Opus 4.7 row → (1000*5 + 100*25)/1e6 = 0.0075.
        self._ingest_planner_event("T-opus", "claude-opus-4-7", 1000, 100, 0, 0)
        self.assertAlmostEqual(self._cost("T-opus"), 0.0075, places=8)

    def test_haiku_row_cost(self) -> None:
        # Q8.3: Haiku 4.5 row → (1000*1 + 100*5)/1e6 = 0.0015.
        self._ingest_planner_event(
            "T-haiku", "claude-haiku-4-5-20251001", 1000, 100, 0, 0)
        self.assertAlmostEqual(self._cost("T-haiku"), 0.0015, places=8)

    def test_unknown_model_falls_back_to_opus(self) -> None:
        # Q8.4: unknown model → Opus 4.7 fallback (0.0075), and is surfaced.
        self._ingest_planner_event("T-future", "claude-future-99", 1000, 100, 0, 0)
        self.assertAlmostEqual(self._cost("T-future"), 0.0075, places=8)
        self.assertIn("claude-future-99",
                      harness_v2.unknown_cost_models(self.con))

    def test_mixed_model_spend_ledger_reconciles(self) -> None:
        # Q8.5: a mixed-model run (Haiku Stage A + Opus Stage B) — the stored
        # spend_ledger snapshot equals the recomputed per_run_summary total.
        run_id = "T-mixed"
        run_dir = self.tmp_path / run_id
        run_dir.mkdir(exist_ok=True)
        evs = [
            _event_row(0, "run.start", {}, run_id=run_id),
            _event_row(10, "planner.stage_a.api_end",
                       {"model": "claude-haiku-4-5-20251001", "input_tokens": 3800,
                        "output_tokens": 50, "cache_creation_input_tokens": 0,
                        "cache_read_input_tokens": 0, "ok": True},
                       run_id=run_id, step_idx=0),
            _event_row(20, "planner.stage_b.api_end",
                       {"model": "claude-opus-4-7", "input_tokens": 900,
                        "output_tokens": 380, "cache_creation_input_tokens": 0,
                        "cache_read_input_tokens": 3479, "ok": True},
                       run_id=run_id, step_idx=0),
            _event_row(30, "run.end", {}, run_id=run_id),
        ]
        (run_dir / "events.jsonl").write_text(
            "\n".join(json.dumps(e) for e in evs) + "\n", encoding="utf-8")
        (run_dir / "mode.txt").write_text("real\n", encoding="utf-8")
        harness_v2.ingest(self.con, run_dir)
        ledger = self.con.execute(
            "SELECT total_cost_usd FROM spend_ledger "
            "WHERE run_id = ? AND superseded = FALSE", [run_id]).fetchone()[0]
        per_run = self.con.execute(
            "SELECT total_cost_usd FROM per_run_summary WHERE run_id = ?",
            [run_id]).fetchone()[0]
        self.assertAlmostEqual(ledger, per_run, places=8)
        # Haiku Stage A (3800*1 + 50*5)/1e6 + Opus Stage B (900*5 + 380*25 + 3479*0.5)/1e6
        expected = (3800 * 1 + 50 * 5) / 1e6 + (900 * 5 + 380 * 25 + 3479 * 0.50) / 1e6
        self.assertAlmostEqual(per_run, expected, places=8)

    def test_phase1c_step4_db_recompute_in_band(self) -> None:
        # Q8.6 (load-bearing): the existing Phase 1c Step 4 exit-sweep, recomputed
        # under honest per-model rates, lands in the Option-2 band $1.00-1.10.
        # state/ is gitignored → skip when the transient sweep DB is absent.
        db = (Path(__file__).resolve().parent.parent
              / "state" / "v3-phase-1c" / "exit-sweep.duckdb")
        if not db.is_file():
            self.skipTest("Phase 1c Step 4 exit-sweep DB not present (state/ gitignored)")
        con = harness_v2.open_db(db)  # CREATE OR REPLACE recomputes the view
        try:
            total = con.execute(
                "SELECT SUM(cost_usd) FROM operations WHERE mode = 'real'"
            ).fetchone()[0]
        finally:
            con.close()
        self.assertGreater(total, 1.00)
        self.assertLess(total, 1.10)


if __name__ == "__main__":
    unittest.main()
