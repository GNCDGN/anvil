"""v2 Phase 1 Step 4 — harness_v2: events.jsonl → DuckDB → views/XLSX.

Consumer of the event stream wired by Steps 1–3. Reads
`<ANVIL_ROOT>/state/runs/<run_id>/events.jsonl`, ingests into the
`events` table at the default DuckDB (v2 Phase 3 Step 1: now
`<ANVIL_ROOT>/state/v2-phase-2/calibration.duckdb` — see `db_path()`;
the pre-`mode` v1 DB is quarantined to `calibration-archived.duckdb`),
exposes the views (operations / per_run_summary / per_task_comparison /
validation_failure_episodes / cumulative_spend_by_task) and an openpyxl
XLSX export. The DB is overridable per invocation via `--db-path`.

The legacy `tools/exam_harness.py` is unchanged and stays around for
backward-compat parsers of the `[planner]` log line. harness_v2.py
reads JSONL only — no log-line parsing.

CLI surface:
    python tools/harness_v2.py ingest <run-dir>
    python tools/harness_v2.py ingest-all [--state-root <path>]
    python tools/harness_v2.py operations [<run-id>]
    python tools/harness_v2.py per-run-summary [--run-id <id>]
    python tools/harness_v2.py per-task-comparison
    python tools/harness_v2.py validation-failure-episodes
    python tools/harness_v2.py export-xlsx <out.xlsx>
    python tools/harness_v2.py --self-check

The `_real_write = Path.write_text` capture at module top mirrors
`anvil/events.py` and `anvil/vault_ops.py`. Tests patch via this seam
when they want to control filesystem behaviour deterministically.
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

import duckdb
from openpyxl import Workbook
from openpyxl.styles import Font

# v3 Phase 2a Step 1 (V3P2A-1): the per-model Planner system-prompt token
# mapping lives canonically in anvil.events; the cache_diagnostics view's
# per-model uncached_user_prompt_equiv CASE generates from it (single source
# of truth — replacing the prior hardcoded `- 3479` literal). harness_v2 runs
# both as a module (tests) and as a script (`python tools/harness_v2.py`); the
# script path puts tools/ on sys.path[0], not the repo root, so anvil isn't
# importable without this shim — the same idiom calibration_runner.py uses.
_REPO_ROOT = Path(__file__).resolve().parent.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))
from anvil.events import PLANNER_SYSTEM_PROMPT_TOKENS_BY_MODEL  # noqa: E402

# Module-scope capture — tests patch via this seam for failure injection.
_real_write = Path.write_text

# ---------------------------------------------------------------------------
# Cost rates — per-model, USD per million tokens. Verified against Anthropic's
# pricing page (platform.claude.com/docs/en/docs/about-claude/pricing,
# fetched 2026-05-26). v3 Phase 1c Step 3.5 (Step3.5C-F1): replaces the prior
# single rate table ($15/$75/$18.75/$1.50) — those were Opus 4.1's rates,
# STALE for the live Opus 4.7 planner model ($5/$25/$6.25/$0.50), and were
# also (wrongly) applied to the Haiku 4.5 canary. Cache: 5m-write = 1.25x
# input, read = 0.1x input. Mirrored in tools/exam_harness.py via import so
# the two harnesses can't drift. Unknown models fall back to Opus 4.7 rates
# (conservative overcharge); unknown_cost_models() surfaces any for
# explicit registration (a Phase 2 new-model addition shows up there).
# ---------------------------------------------------------------------------

MODEL_RATES: dict[str, dict[str, float]] = {
    "claude-opus-4-7": {
        "input": 5.0, "output": 25.0, "cache_create": 6.25, "cache_read": 0.50},
    "claude-haiku-4-5-20251001": {
        "input": 1.0, "output": 5.0, "cache_create": 1.25, "cache_read": 0.10},
}
DEFAULT_MODEL_RATES = MODEL_RATES["claude-opus-4-7"]


def _rate_case(component: str) -> str:
    """Build a SQL CASE mapping the event's model → its per-Mtok rate for
    `component` (input/output/cache_create/cache_read), defaulting to Opus 4.7
    for unknown models. Single source of truth: MODEL_RATES."""
    whens = " ".join(
        f"WHEN '{model}' THEN {rates[component]}"
        for model, rates in MODEL_RATES.items()
    )
    return (f"CASE json_extract_string(e.data, '$.model') "
            f"{whens} ELSE {DEFAULT_MODEL_RATES[component]} END")


def _cost_usd_case_sql() -> str:
    """The cost_usd CASE for the operations view: Coder uses its reported
    CLI cost; null-input rows are 0; planner rows use the per-model token
    formula (rates from MODEL_RATES via _rate_case)."""
    return f"""CASE
        WHEN e.kind LIKE 'coder.%' THEN
            COALESCE(CAST(json_extract(e.data, '$.total_cost_usd') AS DOUBLE), 0.0)
        WHEN CAST(json_extract(e.data, '$.input_tokens') AS BIGINT) IS NULL THEN 0.0
        ELSE (
            COALESCE(CAST(json_extract(e.data, '$.input_tokens') AS BIGINT), 0) * ({_rate_case('input')}) +
            COALESCE(CAST(json_extract(e.data, '$.output_tokens') AS BIGINT), 0) * ({_rate_case('output')}) +
            COALESCE(CAST(json_extract(e.data, '$.cache_creation_input_tokens') AS BIGINT), 0) * ({_rate_case('cache_create')}) +
            COALESCE(CAST(json_extract(e.data, '$.cache_read_input_tokens') AS BIGINT), 0) * ({_rate_case('cache_read')})
        ) / 1000000.0
    END"""


def _system_prompt_tokens_case_sql() -> str:
    """Build a SQL CASE mapping the event's model → its Planner system-prompt
    token count, defaulting to Opus for unknown models. Subtracted in the
    cache_diagnostics uncached_user_prompt_equiv. Single source of truth:
    anvil.events.PLANNER_SYSTEM_PROMPT_TOKENS_BY_MODEL (v3 Phase 2a Step 1,
    Q-A5 — the system constant is per-model: Opus 3479, Haiku 2590)."""
    whens = " ".join(
        f"WHEN '{model}' THEN {tokens}"
        for model, tokens in PLANNER_SYSTEM_PROMPT_TOKENS_BY_MODEL.items()
    )
    default = PLANNER_SYSTEM_PROMPT_TOKENS_BY_MODEL["claude-opus-4-7"]
    return (f"CASE json_extract_string(e.data, '$.model') "
            f"{whens} ELSE {default} END")


def unknown_cost_models(con) -> list[str]:
    """Planner-event models not in MODEL_RATES (priced at the Opus-4.7
    fallback). Empty on the v3 corpus; a Phase 2 new model surfaces here."""
    known = "', '".join(MODEL_RATES)
    rows = con.execute(
        "SELECT DISTINCT json_extract_string(data, '$.model') FROM events "
        "WHERE kind LIKE 'planner.stage_%.api_end' "
        f"AND json_extract_string(data, '$.model') NOT IN ('{known}') "
        "AND json_extract_string(data, '$.model') IS NOT NULL"
    ).fetchall()
    return [r[0] for r in rows]

# ---------------------------------------------------------------------------
# FIELD_MAP — operations-view projection generator
#
# Each entry: column_name -> (kinds_pattern, json_path, sql_cast).
# kinds_pattern is a `|`-separated list of event kinds the field applies to.
# json_path is a `$.field` JSON pointer. sql_cast is the DuckDB type the
# json_extract value is cast to. Some columns are derived (cost_usd,
# cache_hit_rate, ts_start, prompt_chars) and live outside this map.
# ---------------------------------------------------------------------------

FIELD_MAP: dict[str, tuple[str, str, str]] = {
    # Token usage (only on api_end events)
    "input_tokens":         ("planner.stage_a.api_end|planner.stage_b.api_end|planner.stage_c.api_end",
                             "$.input_tokens", "BIGINT"),
    "output_tokens":        ("planner.stage_a.api_end|planner.stage_b.api_end|planner.stage_c.api_end",
                             "$.output_tokens", "BIGINT"),
    "cache_creation_tokens": ("planner.stage_a.api_end|planner.stage_b.api_end|planner.stage_c.api_end",
                              "$.cache_creation_input_tokens", "BIGINT"),
    "cache_read_tokens":    ("planner.stage_a.api_end|planner.stage_b.api_end|planner.stage_c.api_end",
                             "$.cache_read_input_tokens", "BIGINT"),
    # Operation outcome
    "model":                ("*.api_end", "$.model", "VARCHAR"),
    "duration_ms":          ("*", "$.duration_ms", "BIGINT"),
    "response_chars":       ("coder.subprocess.end", "$.stdout_chars", "BIGINT"),
    "files_loaded":         ("planner.stage_b.files_loaded", "$.files_loaded_count", "BIGINT"),
    "files_touched_count":  ("coder.scope_verify", "$.files_touched_count", "BIGINT"),
    "exit_code":            ("coder.subprocess.end", "$.exit_code", "BIGINT"),
    "ok":                   ("*.end|*.reply|*.resolved", "$.ok", "BOOLEAN"),
    "escalation_reason":    ("escalation.raised|escalation.resolved|planner.escalate|coder.preflight.escalate",
                             "$.reason", "VARCHAR"),
    "escalation_user_latency_ms": ("escalation.resolved", "$.latency_ms_user", "BIGINT"),
    "retry_attempt":        ("planner.stage_b.*", "$.retry_attempt", "BIGINT"),
    "out_of_scope_count":   ("coder.scope_verify", "$.out_of_scope_count", "BIGINT"),
}

# Kinds that surface as one row each in the operations view. Each is the
# "terminal" event for an operation. .start events become joined columns
# (ts_start, prompt_chars) via correlated subquery.
OPERATION_KINDS: tuple[str, ...] = (
    "planner.stage_a.api_end",
    "planner.stage_b.api_end",
    "planner.stage_c.api_end",
    # v3 Phase 1b Step 3 (V3P1B-3): the canary's parallel Opus baseline call —
    # a cost-bearing operation so its spend is ledgered (the token-formula cost
    # CASE applies automatically, like any non-coder api_end).
    "planner.stage_a.canary_baseline.api_end",
    "planner.validation.pass",
    "planner.validation.fail",
    "planner.retry.end",
    "planner.escalate",
    "coder.subprocess.end",
    "coder.preflight.escalate",
    "coder.scope_verify",
    "smoke.end",
    "git.commit.end",
    "git.push.end",
    "ssh.stage.end",
    "telegram.send.end",
    "telegram.poll.reply",
    "escalation.resolved",
)


# ---------------------------------------------------------------------------
# Database lifecycle
# ---------------------------------------------------------------------------

def _anvil_root() -> Path:
    """Resolve ANVIL_ROOT from env, defaulting to the repo root (parent
    of the `tools/` dir)."""
    default = Path(__file__).resolve().parent.parent
    return Path(os.environ.get("ANVIL_ROOT", str(default))).expanduser()


def db_path() -> Path:
    """Default DuckDB path under `<ANVIL_ROOT>/state/v2-phase-2/`.

    v2 Phase 3 Step 1: re-pointed from v2-phase-1 → v2-phase-2. The v1 DB
    predates the `mode` column and binder-errors under the composite-key
    views (`LEFT JOIN run_metadata USING (run_id, mode)` — `mode` is
    absent on v1's `events`), so the bare CLI defaulting to it always
    failed. The v1 DB is quarantined to `calibration-archived.duckdb`;
    the clean v2 Phase 2 baseline is the new default. Override per
    invocation with the CLI `--db-path` flag.
    """
    return _anvil_root() / "state" / "v2-phase-2" / "calibration.duckdb"


def open_db(path: Path | None = None) -> duckdb.DuckDBPyConnection:
    """Open (or create) the DuckDB file at `path` (or `db_path()`).

    On first open, creates the schema (events table, run_metadata table,
    three views). Subsequent opens are no-ops at the schema level —
    CREATE TABLE IF NOT EXISTS + CREATE OR REPLACE VIEW are both
    idempotent.
    """
    p = path if path is not None else db_path()
    p.parent.mkdir(parents=True, exist_ok=True)
    con = duckdb.connect(str(p))
    _ensure_schema(con)
    return con


def _ensure_schema(con: duckdb.DuckDBPyConnection) -> None:
    # v2 Phase 2 Step 1: `mode` is a first-class column on events (sourced
    # from `<run_dir>/mode.txt` at ingest time, default `'unknown'`). The
    # delete-then-insert idempotency key is the composite (run_id, mode);
    # a mock-then-real ingest of the same task no longer clobbers the
    # mock half. run_metadata's PK is the same composite.
    con.execute(
        """
        CREATE TABLE IF NOT EXISTS events (
            ts          TIMESTAMP,
            run_id      VARCHAR,
            mode        VARCHAR NOT NULL DEFAULT 'unknown',
            step_idx    INTEGER,
            kind        VARCHAR,
            data        JSON,
            elapsed_ms  BIGINT
        )
        """
    )
    con.execute(
        """
        CREATE TABLE IF NOT EXISTS run_metadata (
            run_id        VARCHAR,
            mode          VARCHAR NOT NULL DEFAULT 'unknown',
            task_id       VARCHAR,
            task_label    VARCHAR,
            ingested_at   TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (run_id, mode)
        )
        """
    )
    # v2 Phase 3 Step 1: spend_ledger — append-only spend history
    # (Candidate A). Every ingest of a (run_id, mode) appends a row
    # capturing that ingest's total cost; the prior non-superseded row
    # for the key is flipped to superseded=TRUE (see `ingest()`). The
    # events/run_metadata composite-key DELETE-then-INSERT overwrite is
    # UNCHANGED — the ledger is purely additive, so the V2P2-2
    # (run_id, mode) invariant holds and none of the four existing views
    # are rewritten. ledger_id auto-increments via a sequence (DuckDB has
    # no SERIAL/AUTOINCREMENT). CREATE ... IF NOT EXISTS migrates the
    # existing v2-phase-2 DB on next open without touching its data.
    con.execute("CREATE SEQUENCE IF NOT EXISTS spend_ledger_seq START 1")
    con.execute(
        """
        CREATE TABLE IF NOT EXISTS spend_ledger (
            ledger_id      INTEGER PRIMARY KEY DEFAULT nextval('spend_ledger_seq'),
            run_id         VARCHAR NOT NULL,
            mode           VARCHAR NOT NULL,
            ingest_ts      TIMESTAMP NOT NULL,
            total_cost_usd DOUBLE NOT NULL,
            superseded     BOOLEAN NOT NULL DEFAULT FALSE
        )
        """
    )
    # v3 Phase 0 Step 2 (V3P0-3): shadow_decisions — per Planner call,
    # what a hypothetical shadow router WOULD have decided vs what the
    # code actually did. Append-shaped per (run_id, mode) like events:
    # ingest DELETE-then-INSERTs on the same composite key (V2P2-2). The
    # `policy_version` column carries Step 1's literal placeholder by
    # default so Phase 1's first real policy populates the existing
    # column rather than triggering an ALTER TABLE migration. CREATE ...
    # IF NOT EXISTS migrates an existing DB on next open without touching
    # its data.
    con.execute(
        """
        CREATE TABLE IF NOT EXISTS shadow_decisions (
            run_id                 VARCHAR NOT NULL,
            mode                   VARCHAR NOT NULL,
            step_idx               INTEGER,
            stage                  VARCHAR,
            ts                     TIMESTAMP,
            shadow_route_candidate VARCHAR,
            shadow_decision_basis  JSON,
            actual_route_taken     VARCHAR,
            agreement              BOOLEAN,
            policy_version         VARCHAR NOT NULL DEFAULT 'v3-phase-0-passive'
        )
        """
    )
    con.execute(_OPERATIONS_VIEW_SQL)
    con.execute(_PER_RUN_SUMMARY_VIEW_SQL)
    con.execute(_PER_TASK_COMPARISON_VIEW_SQL)
    # v2 Phase 2 Step 2: registered last because it SELECTs from
    # `operations` (cost_usd source) and `run_metadata` (task_id).
    con.execute(_VALIDATION_FAILURE_EPISODES_VIEW_SQL)
    # v2 Phase 3 Step 1: cumulative_spend_by_task SELECTs from
    # spend_ledger (above) and run_metadata (task_id).
    con.execute(_CUMULATIVE_SPEND_VIEW_SQL)
    # v3 Phase 0 Step 2 (V3P0-3): champion_challenger_comparison SELECTs
    # from shadow_decisions + operations + run_metadata; registered last.
    con.execute(_CHAMPION_CHALLENGER_VIEW_SQL)
    # v3 Phase 0 Step 3 (V3P0-4): silent_miss_episodes reads the events
    # table + run_metadata; empty-set by construction in Phase 0.
    con.execute(_SILENT_MISS_EPISODES_VIEW_SQL)
    # v3 Phase 0 Step 4 (V3P0-6): cache_diagnostics reads the events
    # table + run_metadata; the three cache-family lines per Planner stage.
    con.execute(_CACHE_DIAGNOSTICS_VIEW_SQL)
    # v3 Phase 2a Step 2 (V3P2A-2): stage_a_selections exposes the recorded
    # selection list + raw response for the Phase 2c rich-context comparator.
    con.execute(_STAGE_A_SELECTIONS_VIEW_SQL)


# ---------------------------------------------------------------------------
# Views
# ---------------------------------------------------------------------------

# operations view: one row per cost-bearing operation across all runs.
# JSON extraction is verbose but explicit — each column documents its
# source kind set + json path. The .start lookup for ts_start /
# prompt_chars is a correlated subquery; works at fixture scale and at
# the calibration sweep's ~hundreds-of-events scale.
_OPERATIONS_VIEW_SQL = """
CREATE OR REPLACE VIEW operations AS
SELECT
    e.run_id,
    rm.task_id,
    rm.task_label,
    rm.mode,
    e.step_idx,
    e.kind AS operation_kind,
    (
        SELECT s.ts FROM events s
        WHERE s.run_id = e.run_id
          AND COALESCE(s.step_idx, -1) = COALESCE(e.step_idx, -1)
          AND s.kind = CASE
            WHEN e.kind LIKE '%.api_end'           THEN REPLACE(e.kind, '.api_end', '.api_start')
            WHEN e.kind = 'coder.subprocess.end'    THEN 'coder.subprocess.start'
            WHEN e.kind = 'smoke.end'               THEN 'smoke.start'
            WHEN e.kind = 'git.commit.end'          THEN 'git.commit.start'
            WHEN e.kind = 'git.push.end'            THEN 'git.push.start'
            WHEN e.kind = 'ssh.stage.end'           THEN 'ssh.stage.start'
            WHEN e.kind = 'telegram.send.end'       THEN 'telegram.send.start'
            WHEN e.kind = 'telegram.poll.reply'     THEN 'telegram.poll.start'
            WHEN e.kind = 'escalation.resolved'     THEN 'escalation.raised'
            ELSE NULL
          END
        ORDER BY s.ts DESC LIMIT 1
    ) AS ts_start,
    e.ts AS ts_end,
    CAST(json_extract(e.data, '$.duration_ms') AS BIGINT) AS duration_ms,
    json_extract_string(e.data, '$.model') AS model,
    CAST(json_extract(e.data, '$.input_tokens') AS BIGINT) AS input_tokens,
    CAST(json_extract(e.data, '$.output_tokens') AS BIGINT) AS output_tokens,
    CAST(json_extract(e.data, '$.cache_creation_input_tokens') AS BIGINT) AS cache_creation_tokens,
    CAST(json_extract(e.data, '$.cache_read_input_tokens') AS BIGINT) AS cache_read_tokens,
    CASE
        WHEN CAST(json_extract(e.data, '$.input_tokens') AS BIGINT) IS NULL THEN NULL
        WHEN (
            COALESCE(CAST(json_extract(e.data, '$.cache_read_input_tokens') AS BIGINT), 0) +
            COALESCE(CAST(json_extract(e.data, '$.cache_creation_input_tokens') AS BIGINT), 0) +
            CAST(json_extract(e.data, '$.input_tokens') AS BIGINT)
        ) = 0 THEN NULL
        ELSE
            COALESCE(CAST(json_extract(e.data, '$.cache_read_input_tokens') AS BIGINT), 0) * 1.0 /
            (
                COALESCE(CAST(json_extract(e.data, '$.cache_read_input_tokens') AS BIGINT), 0) +
                COALESCE(CAST(json_extract(e.data, '$.cache_creation_input_tokens') AS BIGINT), 0) +
                CAST(json_extract(e.data, '$.input_tokens') AS BIGINT)
            )
    END AS cache_hit_rate,
    -- v2 Phase 5 Step 1a: Coder cost is the CLI's *reported* total_cost_usd
    -- (`claude --output-format json`), NOT the token formula. v3 Phase 1c
    -- Step 3.5 (Step3.5C-F1): planner rows now use per-model rates (MODEL_RATES,
    -- via _cost_usd_case_sql) instead of a single Opus-4.1 table — Opus 4.7 and
    -- the Haiku canary are priced honestly. Injected below to keep MODEL_RATES
    -- the single source of truth.
    __COST_USD_CASE__ AS cost_usd,
    (
        SELECT CAST(json_extract(s.data, '$.prompt_chars') AS BIGINT) FROM events s
        WHERE s.run_id = e.run_id
          AND COALESCE(s.step_idx, -1) = COALESCE(e.step_idx, -1)
          AND s.kind = CASE
            WHEN e.kind LIKE '%.api_end'           THEN REPLACE(e.kind, '.api_end', '.api_start')
            WHEN e.kind = 'coder.subprocess.end'    THEN 'coder.subprocess.start'
            ELSE NULL
          END
        ORDER BY s.ts DESC LIMIT 1
    ) AS prompt_chars,
    CAST(json_extract(e.data, '$.stdout_chars') AS BIGINT) AS response_chars,
    CAST(json_extract(e.data, '$.files_loaded_count') AS BIGINT) AS files_loaded,
    CAST(json_extract(e.data, '$.files_touched_count') AS BIGINT) AS files_touched_count,
    CAST(json_extract(e.data, '$.exit_code') AS BIGINT) AS exit_code,
    CAST(json_extract(e.data, '$.ok') AS BOOLEAN) AS ok,
    CASE
        WHEN e.kind = 'planner.validation.pass' THEN 'pass'
        WHEN e.kind = 'planner.validation.fail' THEN 'fail'
        ELSE NULL
    END AS validation_result,
    json_extract_string(e.data, '$.reason') AS escalation_reason,
    CAST(json_extract(e.data, '$.latency_ms_user') AS BIGINT) AS escalation_user_latency_ms,
    CAST(json_extract(e.data, '$.retry_attempt') AS BIGINT) AS retry_attempt,
    CAST(json_extract(e.data, '$.out_of_scope_count') AS BIGINT) AS out_of_scope_count,
    CAST(json_extract(e.data, '$.reply_text_chars') AS BIGINT) AS poll_reply_chars,
    -- v3 Phase 0 Step 1 (V3P0-1): routing observability columns. Populated
    -- on the four model-call kinds (planner.stage_{a,b,c}.api_end,
    -- coder.subprocess.end); NULL on the other operation kinds (additive,
    -- back-compatible). features_seen is kept as DuckDB JSON via
    -- json_extract (NOT json_extract_string — that would stringify it).
    json_extract_string(e.data, '$.route_candidate') AS route_candidate,
    json_extract_string(e.data, '$.route_actual') AS route_actual,
    CAST(json_extract(e.data, '$.route_fallback_fired') AS BOOLEAN) AS route_fallback_fired,
    json_extract_string(e.data, '$.policy_version') AS policy_version,
    json_extract(e.data, '$.features_seen') AS features_seen
FROM events e
-- v2 Phase 2 Step 1: JOIN on the composite (run_id, mode) to match
-- run_metadata's composite PK. With the same run_id potentially
-- present under both 'mock' and 'real', a run_id-only JOIN would
-- duplicate every event row.
LEFT JOIN run_metadata rm USING (run_id, mode)
WHERE e.kind IN (
    'planner.stage_a.api_end',
    'planner.stage_b.api_end',
    'planner.stage_c.api_end',
    'planner.stage_a.canary_baseline.api_end',
    'planner.validation.pass',
    'planner.validation.fail',
    'planner.retry.end',
    'planner.escalate',
    'coder.subprocess.end',
    'coder.preflight.escalate',
    'coder.scope_verify',
    'smoke.end',
    'git.commit.end',
    'git.push.end',
    'ssh.stage.end',
    'telegram.send.end',
    'telegram.poll.reply',
    'escalation.resolved'
)
"""

# v3 Phase 1c Step 3.5: inject the per-model cost CASE (MODEL_RATES is the
# single source of truth; the SQL is generated from it at module load).
_OPERATIONS_VIEW_SQL = _OPERATIONS_VIEW_SQL.replace(
    "__COST_USD_CASE__", _cost_usd_case_sql())

# Column order in operations (mirrored in XLSX export):
_OPERATIONS_COLUMNS: tuple[str, ...] = (
    "run_id", "task_id", "task_label", "mode", "step_idx", "operation_kind",
    "ts_start", "ts_end", "duration_ms", "model",
    "input_tokens", "output_tokens", "cache_creation_tokens", "cache_read_tokens",
    "cache_hit_rate", "cost_usd",
    "prompt_chars", "response_chars",
    "files_loaded", "files_touched_count",
    "exit_code", "ok", "validation_result",
    "escalation_reason", "escalation_user_latency_ms",
    "retry_attempt", "out_of_scope_count", "poll_reply_chars",
    # v3 Phase 0 Step 1 (V3P0-1): routing observability columns.
    "route_candidate", "route_actual", "route_fallback_fired",
    "policy_version", "features_seen",
)


_PER_RUN_SUMMARY_VIEW_SQL = """
CREATE OR REPLACE VIEW per_run_summary AS
SELECT
    e.run_id,
    rm.task_id,
    rm.task_label,
    rm.mode,
    (SELECT COALESCE(SUM(cost_usd), 0.0) FROM operations o WHERE o.run_id = e.run_id) AS total_cost_usd,
    EXTRACT(EPOCH FROM (MAX(e.ts) - MIN(e.ts))) AS total_duration_s,
    COUNT(*) FILTER (WHERE e.kind LIKE 'planner.stage_%' AND e.kind LIKE '%.api_end') AS planner_calls,
    COUNT(*) FILTER (WHERE e.kind = 'coder.subprocess.end') AS coder_calls,
    COUNT(*) FILTER (WHERE e.kind = 'escalation.raised') AS escalations,
    BOOL_OR(e.kind = 'run.resume') AS resumed,
    (
        SELECT e2.kind FROM events e2
        WHERE e2.run_id = e.run_id ORDER BY e2.ts DESC LIMIT 1
    ) AS terminal_event
FROM events e
-- v2 Phase 2 Step 1: composite-key JOIN; see operations view comment.
LEFT JOIN run_metadata rm USING (run_id, mode)
GROUP BY e.run_id, rm.task_id, rm.task_label, rm.mode
"""

_PER_RUN_SUMMARY_COLUMNS: tuple[str, ...] = (
    "run_id", "task_id", "task_label", "mode",
    "total_cost_usd", "total_duration_s",
    "planner_calls", "coder_calls", "escalations",
    "resumed", "terminal_event",
)


# v2 Phase 2 Step 1: rewrite per_task_comparison as a FULL OUTER JOIN
# between the mock half and the real half of `per_run_summary`, keyed on
# task_id. Each half is a WHERE-mode-filtered subquery — explicit about
# which mode column each metric is sourced from. The prior CASE-pivot
# version was functionally equivalent under v2 Phase 1's run_id-unique
# constraint; under v2 Phase 2's composite key, T1 mock and T1 real are
# two rows sharing task_id, and the JOIN shape makes the cross-mode read
# unambiguous (no MAX() across the two rows).
_PER_TASK_COMPARISON_VIEW_SQL = """
CREATE OR REPLACE VIEW per_task_comparison AS
WITH
    mock_runs AS (
        SELECT task_id, planner_calls, total_duration_s, total_cost_usd
        FROM per_run_summary
        WHERE mode = 'mock' AND task_id IS NOT NULL
    ),
    real_runs AS (
        SELECT task_id, planner_calls, total_duration_s, total_cost_usd
        FROM per_run_summary
        WHERE mode = 'real' AND task_id IS NOT NULL
    )
SELECT
    COALESCE(m.task_id, r.task_id) AS task_id,
    m.planner_calls   AS planner_calls_mock,
    r.planner_calls   AS planner_calls_real,
    m.total_duration_s AS total_duration_mock,
    r.total_duration_s AS total_duration_real,
    r.total_cost_usd  AS total_cost_real,
    m.total_duration_s AS framework_overhead_s
FROM mock_runs m
FULL OUTER JOIN real_runs r USING (task_id)
"""

_PER_TASK_COMPARISON_COLUMNS: tuple[str, ...] = (
    "task_id",
    "planner_calls_mock", "planner_calls_real",
    "total_duration_mock", "total_duration_real",
    "total_cost_real",
    "framework_overhead_s",
)


# v2 Phase 2 Step 2: validation_failure_episodes
#
# One row per (run_id, mode, step_idx) that had at least one
# `planner.validation.fail` event. The shape answers "how much did
# validation failure cost on this run, and did it recover or escalate?"
# in one SELECT — which the v2 Phase 1 exam (Q3) and the v2 Phase 2 exam
# (Q2) both want to grade against the calibration sweep.
#
# Episode terminator simplification (per the brief's carve-out): rather
# than encoding "a stage_b.api_end exists after the last fail with no
# escalate in between", we encode the inverse — `recovered = NOT
# EXISTS(planner.escalate on (run_id, step_idx))`. These are logically
# equivalent for the two-attempt retry pattern in `_run_stage_b_with_retry`
# (the only Stage-B retry shape in production); the brief explicitly
# permits the swap. Comment kept in the view body so future me sees
# the call.
#
# cost_usd source: the production cost formula lives in the `operations`
# view (token-weighted CASE). To avoid duplicating the formula, this
# view SELECTs `operations.cost_usd` for `operation_kind =
# 'planner.stage_b.api_end'`. Any future cost-formula change in
# `operations` flows through automatically.
#
# first_error / second_error come straight from `events.data` on the
# `planner.validation.fail` rows; the field name is `first_error` per
# planner.py:382 (and `:471` for the retry's fail). Note: each fail
# event records its OWN error under the `first_error` key — the names
# `first_error` / `second_error` on this view refer to the chronological
# 1st/2nd validation failure within the episode, not the JSON field
# name on the source event.
_VALIDATION_FAILURE_EPISODES_VIEW_SQL = """
CREATE OR REPLACE VIEW validation_failure_episodes AS
WITH
    fails_numbered AS (
        SELECT
            e.run_id,
            e.mode,
            e.step_idx,
            e.ts,
            json_extract_string(e.data, '$.first_error') AS error_text,
            ROW_NUMBER() OVER (
                PARTITION BY e.run_id, e.mode, e.step_idx ORDER BY e.ts
            ) AS rn
        FROM events e
        WHERE e.kind = 'planner.validation.fail'
          AND e.step_idx IS NOT NULL
    ),
    fail_metrics AS (
        SELECT
            run_id,
            mode,
            step_idx,
            COUNT(*) AS n_validation_fails,
            MAX(CASE WHEN rn = 1 THEN error_text END) AS first_error,
            MAX(CASE WHEN rn = 2 THEN error_text END) AS second_error
        FROM fails_numbered
        GROUP BY run_id, mode, step_idx
    ),
    stage_b_calls AS (
        SELECT
            run_id,
            mode,
            step_idx,
            cost_usd,
            ts_end,
            ROW_NUMBER() OVER (
                PARTITION BY run_id, mode, step_idx ORDER BY ts_end
            ) AS rn
        FROM operations
        WHERE operation_kind = 'planner.stage_b.api_end'
          AND step_idx IS NOT NULL
    ),
    api_metrics AS (
        SELECT
            run_id,
            mode,
            step_idx,
            COUNT(*) FILTER (WHERE rn > 1) AS extra_api_calls,
            COALESCE(
                SUM(cost_usd) FILTER (WHERE rn > 1), 0.0
            ) AS total_extra_cost_usd
        FROM stage_b_calls
        GROUP BY run_id, mode, step_idx
    ),
    escalations AS (
        SELECT DISTINCT run_id, mode, step_idx
        FROM events
        WHERE kind = 'planner.escalate'
          AND step_idx IS NOT NULL
    )
SELECT
    f.run_id,
    rm.task_id,
    f.mode,
    f.step_idx,
    f.n_validation_fails,
    f.first_error,
    f.second_error,
    (esc.run_id IS NULL) AS recovered,
    COALESCE(am.extra_api_calls, 0)        AS extra_api_calls,
    COALESCE(am.total_extra_cost_usd, 0.0) AS total_extra_cost_usd
FROM fail_metrics f
LEFT JOIN run_metadata rm USING (run_id, mode)
LEFT JOIN api_metrics am  USING (run_id, mode, step_idx)
LEFT JOIN escalations esc USING (run_id, mode, step_idx)
"""

_VALIDATION_FAILURE_EPISODES_COLUMNS: tuple[str, ...] = (
    "run_id", "task_id", "mode", "step_idx",
    "n_validation_fails",
    "first_error", "second_error",
    "recovered",
    "extra_api_calls", "total_extra_cost_usd",
)


# v2 Phase 3 Step 1: cumulative_spend_by_task — actual-spend
# reconciliation surface (Candidate A).
#
# One row per (task_id, run_id, mode). `n_ingests` is the spend_ledger
# history depth for the key; `cumulative_cost_usd` sums every ingest —
# the ACTUAL spend, including superseded re-ingests; `latest_cost_usd`
# is the single non-superseded snapshot (equals
# per_run_summary.total_cost_usd for the key). The gap between
# cumulative and latest is the previously-invisible re-run spend — the
# v2 Phase 2 $0.29 T6-overwrite defect class, now reconcilable.
#
# spend_ledger has N rows per (run_id, mode); run_metadata has exactly
# one (overwrite semantics preserved), so the LEFT JOIN is 1:1 from the
# ledger side — no fan-out. GROUP BY collapses to one row per key.
#
# Reconciliation arithmetic (runnable as single queries):
#   actual_real_spend = SUM(total_cost_usd) FROM spend_ledger WHERE mode='real'
#   recorded_latest   = SUM(total_cost_usd) FROM per_run_summary WHERE mode='real'
#   actual - recorded = SUM of superseded ledger rows (the invisible re-run spend)
_CUMULATIVE_SPEND_VIEW_SQL = """
CREATE OR REPLACE VIEW cumulative_spend_by_task AS
SELECT
    rm.task_id,
    sl.run_id,
    sl.mode,
    COUNT(*) AS n_ingests,
    SUM(sl.total_cost_usd) AS cumulative_cost_usd,
    MAX(sl.total_cost_usd) FILTER (WHERE sl.superseded = FALSE) AS latest_cost_usd
FROM spend_ledger sl
LEFT JOIN run_metadata rm USING (run_id, mode)
GROUP BY rm.task_id, sl.run_id, sl.mode
"""

_CUMULATIVE_SPEND_COLUMNS: tuple[str, ...] = (
    "task_id", "run_id", "mode",
    "n_ingests", "cumulative_cost_usd", "latest_cost_usd",
)


# v3 Phase 0 Step 2 (V3P0-3): champion_challenger_comparison.
#
# Per (stage, policy_version, task_id): how often the shadow router
# agreed with the actual route, disagreed, and how many fallbacks fired.
# Phase 0's placeholder shadow rule always picks Opus and the actual
# Planner route is always Opus, so agreement_count == total and
# disagreement_count == 0 on every row; fallback_fired_count is always 0
# (no fallback paths exist). Phase 1's first real Stage A rule produces
# the first non-zero disagreement/fallback delta against this baseline.
#
# Built as two independently-aggregated CTEs joined on (stage, task_id)
# rather than a row-level join, so a Stage B retry (two stage_b.api_end
# rows + two paired shadow.decision rows at one step_idx) does NOT
# fan-out the counts — each CTE aggregates its own table directly.
# task_id is derived via the run_metadata JOIN (mirrors the operations
# view's pattern) rather than an inline derive_task regex, so the regex
# stays in one place. The view carries task_id so it is joinable to
# per_task_comparison.
#
# ops_agg's stage comes from features_seen.stage (the same "A"/"B"/"C"
# letters shadow_decisions.stage carries); the WHERE filters to Planner
# api_end rows so Coder rows (stage="coder") don't enter the comparison.
_CHAMPION_CHALLENGER_VIEW_SQL = """
CREATE OR REPLACE VIEW champion_challenger_comparison AS
WITH shadow_agg AS (
    SELECT
        sd.stage AS stage,
        rm.task_id AS task_id,
        COUNT(*) FILTER (WHERE sd.agreement = TRUE)  AS agreement_count,
        COUNT(*) FILTER (WHERE sd.agreement = FALSE) AS disagreement_count
    FROM shadow_decisions sd
    LEFT JOIN run_metadata rm USING (run_id, mode)
    GROUP BY sd.stage, rm.task_id
),
ops_agg AS (
    SELECT
        json_extract_string(features_seen, '$.stage') AS stage,
        task_id,
        ANY_VALUE(policy_version) AS policy_version,
        COUNT(*) FILTER (WHERE route_fallback_fired = TRUE) AS fallback_fired_count
    FROM operations
    WHERE operation_kind LIKE 'planner.stage_%.api_end'
    GROUP BY json_extract_string(features_seen, '$.stage'), task_id
)
SELECT
    s.stage,
    o.policy_version,
    s.task_id,
    s.agreement_count,
    s.disagreement_count,
    COALESCE(o.fallback_fired_count, 0) AS fallback_fired_count
FROM shadow_agg s
LEFT JOIN ops_agg o ON s.stage = o.stage AND s.task_id = o.task_id
"""

_CHAMPION_CHALLENGER_COLUMNS: tuple[str, ...] = (
    "stage", "policy_version", "task_id",
    "agreement_count", "disagreement_count", "fallback_fired_count",
)


# v3 Phase 0 Step 3 (V3P0-4): silent_miss_episodes.
#
# Per-(run_id, mode, step_idx) episode for each stage_a.shadow_compare.end
# that recorded a non-zero silent_miss_count — i.e. a Stage A call where
# the routed selection dropped context the baseline kept. The
# `WHERE silent_miss_count > 0` filter makes this an EMPTY-SET query in
# Phase 0 by construction (routed == baseline → silent_miss_count always
# 0), and it materializes the moment Phase 1 feeds a real cheap-vs-Opus
# baseline. Reads the events table directly (no new table); events are
# already (run_id, mode) composite-keyed, so V2P2-2 holds without an
# ingest change. task_id via the run_metadata join (operations-view
# pattern; no inline derive_task regex).
_SILENT_MISS_EPISODES_VIEW_SQL = """
CREATE OR REPLACE VIEW silent_miss_episodes AS
SELECT
    e.run_id,
    rm.task_id,
    e.mode,
    e.step_idx,
    e.ts,
    CAST(json_extract(e.data, '$.silent_miss_count') AS BIGINT) AS silent_miss_count,
    CAST(json_extract(e.data, '$.hallucination_count') AS BIGINT) AS hallucination_count,
    CAST(json_extract(e.data, '$.jaccard_similarity') AS DOUBLE) AS jaccard_similarity
FROM events e
LEFT JOIN run_metadata rm USING (run_id, mode)
WHERE e.kind = 'stage_a.shadow_compare.end'
  AND CAST(json_extract(e.data, '$.silent_miss_count') AS BIGINT) > 0
"""

_SILENT_MISS_EPISODES_COLUMNS: tuple[str, ...] = (
    "run_id", "task_id", "mode", "step_idx", "ts",
    "silent_miss_count", "hallucination_count", "jaccard_similarity",
)


# v3 Phase 0 Step 4 (V3P0-6): cache_diagnostics.
#
# The three cache-family telemetry lines per Planner stage event, exposed
# per-stage and per-mode. Reads the events table directly (the three
# fields live in the planner.stage_*.api_end data JSON; no operations-view
# column needed). vault_index_hit is BOOLEAN (null on Stage B/C by
# construction — Q(c)); candidate_user_block_sizes stays as DuckDB JSON;
# seconds_since_cache_creation is DOUBLE (null on cache_creation calls and
# before any creation). candidate_block_sizes_sum sums the four canonical
# blocks. v3 Phase 1c Step 1 (Step1C-F1): the criterion-3 sum-check is the
# AFFINE relation block_sum × block_token_inflation_factor(model) +
# PLANNER_USER_TEMPLATE_TOKENS (407) ≈ uncached_user_prompt_equiv, where the
# equiv column (added here) is cache-invariant: input_tokens + cache_read +
# cache_creation − the per-model system prompt. v3 Phase 2a Step 1 (V3P2A-1,
# Q-A5): the system constant is PER-MODEL (Opus 3479, Haiku 2590) via an
# injected CASE — the `model` column lets consumers grade BOTH models at their
# native slope (Opus 1.64, Haiku 1.18); the Phase 1c Opus-only filter is
# retired. task_id via the run_metadata join (operations-view pattern).
_CACHE_DIAGNOSTICS_VIEW_SQL = """
CREATE OR REPLACE VIEW cache_diagnostics AS
SELECT
    e.run_id,
    rm.task_id,
    e.mode,
    e.step_idx,
    CASE e.kind
        WHEN 'planner.stage_a.api_end' THEN 'A'
        WHEN 'planner.stage_b.api_end' THEN 'B'
        WHEN 'planner.stage_c.api_end' THEN 'C'
    END AS stage,
    CAST(json_extract(e.data, '$.vault_index_hit') AS BOOLEAN) AS vault_index_hit,
    json_extract(e.data, '$.candidate_user_block_sizes') AS candidate_user_block_sizes,
    CAST(json_extract(e.data, '$.seconds_since_cache_creation') AS DOUBLE)
        AS seconds_since_cache_creation,
    (
        COALESCE(CAST(json_extract(e.data, '$.candidate_user_block_sizes.brief') AS BIGINT), 0)
      + COALESCE(CAST(json_extract(e.data, '$.candidate_user_block_sizes.state') AS BIGINT), 0)
      + COALESCE(CAST(json_extract(e.data, '$.candidate_user_block_sizes.vault_files') AS BIGINT), 0)
      + COALESCE(CAST(json_extract(e.data, '$.candidate_user_block_sizes.prior_step') AS BIGINT), 0)
    ) AS candidate_block_sizes_sum,
    CAST(json_extract(e.data, '$.input_tokens') AS BIGINT) AS input_tokens,
    json_extract_string(e.data, '$.model') AS model,
    CAST(json_extract(e.data, '$.cache_read_input_tokens') AS BIGINT)
        AS cache_read_input_tokens,
    CAST(json_extract(e.data, '$.cache_creation_input_tokens') AS BIGINT)
        AS cache_creation_input_tokens,
    -- v3 Phase 1c Step 1 (Step1C-F1): cache-invariant uncached-user-prompt
    -- token equivalent. The Planner system prompt moves between input_tokens
    -- and cache_read/creation, so adding them back and subtracting it isolates
    -- the user-prompt total. v3 Phase 2a Step 1 (V3P2A-1, Q-A5): the system
    -- constant is PER-MODEL (Opus 3479, Haiku 2590) via the injected CASE
    -- below — so the equiv is honest for BOTH models, not Opus-only. The CASE
    -- generates from anvil.events.PLANNER_SYSTEM_PROMPT_TOKENS_BY_MODEL.
    (
        CAST(json_extract(e.data, '$.input_tokens') AS BIGINT)
      + COALESCE(CAST(json_extract(e.data, '$.cache_read_input_tokens') AS BIGINT), 0)
      + COALESCE(CAST(json_extract(e.data, '$.cache_creation_input_tokens') AS BIGINT), 0)
      - (__SYSTEM_PROMPT_TOKENS_CASE__)
    ) AS uncached_user_prompt_equiv
FROM events e
LEFT JOIN run_metadata rm USING (run_id, mode)
WHERE e.kind IN (
    'planner.stage_a.api_end',
    'planner.stage_b.api_end',
    'planner.stage_c.api_end'
)
"""

# v3 Phase 2a Step 1 (V3P2A-1): inject the per-model system-prompt CASE
# (anvil.events.PLANNER_SYSTEM_PROMPT_TOKENS_BY_MODEL is the single source of
# truth; mirrors the __COST_USD_CASE__ injection for the operations view).
_CACHE_DIAGNOSTICS_VIEW_SQL = _CACHE_DIAGNOSTICS_VIEW_SQL.replace(
    "__SYSTEM_PROMPT_TOKENS_CASE__", _system_prompt_tokens_case_sql())

_CACHE_DIAGNOSTICS_COLUMNS: tuple[str, ...] = (
    "run_id", "task_id", "mode", "step_idx", "stage",
    "vault_index_hit", "candidate_user_block_sizes",
    "seconds_since_cache_creation", "candidate_block_sizes_sum",
    "input_tokens", "model", "cache_read_input_tokens",
    "cache_creation_input_tokens", "uncached_user_prompt_equiv",
)


# v3 Phase 2a Step 2 (V3P2A-2): stage_a_selections.
#
# Exposes the Step 2 recording substrate per Stage A parse: the selection LIST
# (selected_paths — the in-index, deduplicated paths Stage B loads, recorded so
# a Phase 2c comparator can grade Haiku vs Opus on selection CONTENT, not just
# count) plus the model's pre-parser response (raw_response_text + truncated).
# These live in the events.data JSON — the schemaless event payload, same store
# as candidate_user_block_sizes — so the ingest path needs no per-field column
# (the round-trip is automatic). This view projects them via json_extract;
# selected_paths round-trips as a JSON array (content + order preserved).
# paths_returned is retained and equals len(selected_paths) (Step 2 invariant).
_STAGE_A_SELECTIONS_VIEW_SQL = """
CREATE OR REPLACE VIEW stage_a_selections AS
SELECT
    e.run_id,
    rm.task_id,
    e.mode,
    e.step_idx,
    CAST(json_extract(e.data, '$.paths_returned') AS BIGINT) AS paths_returned,
    json_extract(e.data, '$.selected_paths') AS selected_paths,
    json_extract_string(e.data, '$.raw_response_text') AS raw_response_text,
    CAST(json_extract(e.data, '$.truncated') AS BOOLEAN) AS truncated,
    CAST(json_extract(e.data, '$.paths_dropped_as_hallucinated') AS BIGINT)
        AS paths_dropped_as_hallucinated
FROM events e
LEFT JOIN run_metadata rm USING (run_id, mode)
WHERE e.kind = 'planner.stage_a.parsed'
"""

_STAGE_A_SELECTIONS_COLUMNS: tuple[str, ...] = (
    "run_id", "task_id", "mode", "step_idx", "paths_returned",
    "selected_paths", "raw_response_text", "truncated",
    "paths_dropped_as_hallucinated",
)


# ---------------------------------------------------------------------------
# Ingest
# ---------------------------------------------------------------------------

_RUN_DIR_RE = re.compile(r"^(T\d+)(?:-(.+))?$")

# v2 Phase 2 Step 1: calibration run_ids carry the mode as a trailing
# segment (`T1-doc-edit-mock`, `T1-doc-edit-real`). The task_label is
# conceptually mode-independent — mode is its own column. Strip the
# suffix before applying the regex so `task_label` stays `doc-edit`,
# `out-of-scope`, etc., not `doc-edit-mock`.
_MODE_SUFFIXES = ("-mock", "-real", "-unknown")


def derive_task(run_id: str) -> tuple[str, str]:
    """Derive (task_id, task_label) from a run-dir name.

    Calibration runs use `T<N>` or `T<N>-<label>` prefixes; v2 Phase 2
    appends a `-mock`/`-real`/`-unknown` mode segment after the label.
    Real ANVIL runs use `<YYYY-MM-DD-HHMM>-<slug>`. For non-calibration
    shapes, task_id falls back to the run_id and task_label to "".
    """
    stripped = run_id
    for suf in _MODE_SUFFIXES:
        if stripped.endswith(suf):
            stripped = stripped[: -len(suf)]
            break
    m = _RUN_DIR_RE.match(stripped)
    if m:
        return m.group(1), (m.group(2) or "")
    return run_id, ""


def _read_mode(run_dir: Path) -> str:
    """Read `<run_dir>/mode.txt` if present; default `"unknown"` otherwise."""
    p = run_dir / "mode.txt"
    if not p.is_file():
        return "unknown"
    try:
        return p.read_text(encoding="utf-8").strip() or "unknown"
    except Exception:  # noqa: BLE001 — never-raise
        return "unknown"


def _parse_jsonl(path: Path) -> list[dict]:
    """Parse a JSONL file into a list of dicts. Silently drops malformed lines."""
    out: list[dict] = []
    if not path.is_file():
        return out
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            out.append(json.loads(line))
        except json.JSONDecodeError:
            # Malformed lines are dropped (consistent with the never-raise
            # contract of the events.py producer side).
            continue
    return out


def ingest(con: duckdb.DuckDBPyConnection, run_dir: Path) -> dict:
    """Ingest one run-dir's events.jsonl into the events table.

    Idempotent: deletes any prior rows for this run_id, then inserts.
    Returns a small summary dict for the caller (event count + dropped
    count + task_id / mode resolution).
    """
    run_dir = Path(run_dir).resolve()
    run_id = run_dir.name
    task_id, task_label = derive_task(run_id)
    mode = _read_mode(run_dir)

    events_path = run_dir / "events.jsonl"
    rows = _parse_jsonl(events_path)

    con.execute("BEGIN")
    try:
        # v2 Phase 2 Step 1: composite (run_id, mode) idempotency key.
        # A mock-then-real ingest of the same task no longer clobbers the
        # mock half because the modes differ. `mode` is stamped on every
        # event row at ingest time (sourced from `<run_dir>/mode.txt`).
        con.execute(
            "DELETE FROM events WHERE run_id = ? AND mode = ?",
            [run_id, mode],
        )
        con.execute(
            "DELETE FROM run_metadata WHERE run_id = ? AND mode = ?",
            [run_id, mode],
        )
        # v3 Phase 0 Step 2 (V3P0-3): shadow_decisions shares the events
        # composite-key DELETE-then-INSERT idempotency (V2P2-2). Deleted
        # here inside the same transaction; rows re-inserted from the
        # shadow.decision events below.
        con.execute(
            "DELETE FROM shadow_decisions WHERE run_id = ? AND mode = ?",
            [run_id, mode],
        )
        con.execute(
            "INSERT INTO run_metadata (run_id, mode, task_id, task_label) "
            "VALUES (?, ?, ?, ?)",
            [run_id, mode, task_id, task_label],
        )
        for ev in rows:
            con.execute(
                """
                INSERT INTO events (ts, run_id, mode, step_idx, kind, data, elapsed_ms)
                VALUES (CAST(? AS TIMESTAMP), ?, ?, ?, ?, CAST(? AS JSON), ?)
                """,
                [
                    ev.get("ts"),
                    ev.get("run_id") or run_id,
                    mode,
                    ev.get("step_idx"),
                    ev.get("kind"),
                    json.dumps(ev.get("data") or {}),
                    int(ev.get("elapsed_ms") or 0),
                ],
            )
        # v3 Phase 0 Step 2 (V3P0-3): populate shadow_decisions from the
        # shadow.decision events in the same JSONL. One table row per
        # event; fields read from the event's `data` payload. step_idx
        # comes from the event top-level (consistent with the events table).
        # v3 Phase 1a Step 3 (V3P1A-3): policy_version is now read from the
        # event data and inserted into the column (criterion 4 — stamp every
        # shadow_decisions row). Phase 0 events that lack the key fall back to
        # the column's default stamp, so the existing self-check fixtures keep
        # producing 'v3-phase-0-passive' rows unchanged.
        for ev in rows:
            if ev.get("kind") != "shadow.decision":
                continue
            d = ev.get("data") or {}
            con.execute(
                """
                INSERT INTO shadow_decisions
                    (run_id, mode, step_idx, stage, ts,
                     shadow_route_candidate, shadow_decision_basis,
                     actual_route_taken, agreement, policy_version)
                VALUES (?, ?, ?, ?, CAST(? AS TIMESTAMP),
                        ?, CAST(? AS JSON), ?, ?, ?)
                """,
                [
                    ev.get("run_id") or run_id,
                    mode,
                    ev.get("step_idx"),
                    d.get("stage"),
                    ev.get("ts"),
                    d.get("shadow_route_candidate"),
                    json.dumps(d.get("shadow_decision_basis") or {}),
                    d.get("actual_route_taken"),
                    d.get("agreement"),
                    d.get("policy_version") or "v3-phase-0-passive",
                ],
            )
        # v2 Phase 3 Step 1: append to spend_ledger inside the SAME
        # transaction as the events overwrite. Three steps, atomic:
        #   1. supersede the prior non-superseded row for (run_id, mode)
        #      (zero rows on a first ingest);
        #   2. snapshot this ingest's total cost from the operations view
        #      — the run_id-only filter mirrors per_run_summary's
        #      total_cost_usd subquery byte-for-byte (run_id is
        #      mode-suffixed in calibration, so it equals the
        #      per-(run_id, mode) sum);
        #   3. append the new current row (superseded=FALSE).
        # A crash before COMMIT rolls back all of it, so the ledger can
        # never hold two non-superseded rows for one key. The cost is
        # computed AFTER the events insert so the operations view sees
        # this ingest's freshly-inserted rows (same-connection MVCC).
        con.execute(
            "UPDATE spend_ledger SET superseded = TRUE "
            "WHERE run_id = ? AND mode = ? AND superseded = FALSE",
            [run_id, mode],
        )
        cost_row = con.execute(
            "SELECT COALESCE(SUM(cost_usd), 0.0) FROM operations WHERE run_id = ?",
            [run_id],
        ).fetchone()
        total_cost_usd = float(cost_row[0] or 0.0)
        con.execute(
            "INSERT INTO spend_ledger "
            "(run_id, mode, ingest_ts, total_cost_usd, superseded) "
            "VALUES (?, ?, CURRENT_TIMESTAMP, ?, FALSE)",
            [run_id, mode, total_cost_usd],
        )
        con.execute("COMMIT")
    except Exception:
        con.execute("ROLLBACK")
        raise

    return {
        "run_id": run_id,
        "task_id": task_id,
        "task_label": task_label,
        "mode": mode,
        "events_ingested": len(rows),
    }


def ingest_all(
    con: duckdb.DuckDBPyConnection, state_root: Path | None = None,
) -> list[dict]:
    """Iterate every subdir under `<state_root>/runs/` (or `<ANVIL_ROOT>/state/runs/`)
    containing an `events.jsonl` and ingest each. Returns a list of
    per-run summary dicts in dir-name order.

    Per Step 3 outcome finding 1: discovers run-dirs by globbing rather
    than requiring an explicit list — every directory with an
    `events.jsonl` is a candidate.
    """
    root = state_root if state_root is not None else (_anvil_root() / "state")
    runs_root = root / "runs"
    if not runs_root.is_dir():
        return []
    out: list[dict] = []
    for run_dir in sorted(p for p in runs_root.iterdir() if p.is_dir()):
        if (run_dir / "events.jsonl").is_file():
            out.append(ingest(con, run_dir))
    return out


# ---------------------------------------------------------------------------
# Query
# ---------------------------------------------------------------------------

def query_operations(
    con: duckdb.DuckDBPyConnection, run_id: str | None = None,
) -> list[tuple]:
    if run_id:
        return con.execute(
            f"SELECT {', '.join(_OPERATIONS_COLUMNS)} FROM operations "
            "WHERE run_id = ? ORDER BY ts_end",
            [run_id],
        ).fetchall()
    return con.execute(
        f"SELECT {', '.join(_OPERATIONS_COLUMNS)} FROM operations "
        "ORDER BY run_id, ts_end"
    ).fetchall()


def query_per_run_summary(
    con: duckdb.DuckDBPyConnection, run_id: str | None = None,
) -> list[tuple]:
    if run_id:
        return con.execute(
            f"SELECT {', '.join(_PER_RUN_SUMMARY_COLUMNS)} FROM per_run_summary "
            "WHERE run_id = ?",
            [run_id],
        ).fetchall()
    return con.execute(
        f"SELECT {', '.join(_PER_RUN_SUMMARY_COLUMNS)} FROM per_run_summary "
        "ORDER BY run_id"
    ).fetchall()


def query_per_task_comparison(con: duckdb.DuckDBPyConnection) -> list[tuple]:
    return con.execute(
        f"SELECT {', '.join(_PER_TASK_COMPARISON_COLUMNS)} FROM per_task_comparison "
        "ORDER BY task_id"
    ).fetchall()


def query_validation_failure_episodes(
    con: duckdb.DuckDBPyConnection,
) -> list[tuple]:
    """v2 Phase 2 Step 2: one row per (run_id, mode, step_idx) episode
    that had at least one `planner.validation.fail`. See the
    `_VALIDATION_FAILURE_EPISODES_VIEW_SQL` docstring for the columns
    and the recovered-vs-escalated simplification rationale."""
    return con.execute(
        f"SELECT {', '.join(_VALIDATION_FAILURE_EPISODES_COLUMNS)} "
        "FROM validation_failure_episodes "
        "ORDER BY run_id, step_idx"
    ).fetchall()


def query_cumulative_spend(
    con: duckdb.DuckDBPyConnection, mode: str | None = None,
) -> list[tuple]:
    """v2 Phase 3 Step 1: cumulative_spend_by_task rows, optionally
    filtered to one mode (`real`/`mock`). Ordered by run_id for stable
    output. See `_CUMULATIVE_SPEND_VIEW_SQL` for the columns and the
    reconciliation arithmetic."""
    cols = ", ".join(_CUMULATIVE_SPEND_COLUMNS)
    if mode:
        return con.execute(
            f"SELECT {cols} FROM cumulative_spend_by_task "
            "WHERE mode = ? ORDER BY run_id",
            [mode],
        ).fetchall()
    return con.execute(
        f"SELECT {cols} FROM cumulative_spend_by_task ORDER BY run_id"
    ).fetchall()


# ---------------------------------------------------------------------------
# XLSX export
# ---------------------------------------------------------------------------

def export_xlsx(con: duckdb.DuckDBPyConnection, out_path: Path) -> None:
    """Export four sheets to `out_path`:
      - operations
      - per_run_summary
      - per_task_comparison
      - validation_episodes (v2 Phase 2 Step 2)

    Header row bold, frozen panes on row 1, no charts."""
    wb = Workbook()
    # Default sheet → operations.
    ws_ops = wb.active
    ws_ops.title = "operations"
    _write_sheet(ws_ops, _OPERATIONS_COLUMNS, query_operations(con),
                 cost_columns={"cost_usd"},
                 rate_columns={"cache_hit_rate"})

    ws_runs = wb.create_sheet("per_run_summary")
    _write_sheet(ws_runs, _PER_RUN_SUMMARY_COLUMNS, query_per_run_summary(con),
                 cost_columns={"total_cost_usd"})

    ws_tasks = wb.create_sheet("per_task_comparison")
    _write_sheet(ws_tasks, _PER_TASK_COMPARISON_COLUMNS,
                 query_per_task_comparison(con),
                 cost_columns={"total_cost_real"})

    # v2 Phase 2 Step 2: validation_episodes sheet. Column order
    # mirrors `_VALIDATION_FAILURE_EPISODES_COLUMNS`; the
    # `total_extra_cost_usd` column gets the `$` number format like
    # the other cost columns elsewhere in the export.
    ws_eps = wb.create_sheet("validation_episodes")
    _write_sheet(ws_eps, _VALIDATION_FAILURE_EPISODES_COLUMNS,
                 query_validation_failure_episodes(con),
                 cost_columns={"total_extra_cost_usd"})

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))


def _write_sheet(
    ws,
    columns: tuple[str, ...],
    rows: list[tuple],
    *,
    cost_columns: set[str] | None = None,
    rate_columns: set[str] | None = None,
) -> None:
    cost_columns = cost_columns or set()
    rate_columns = rate_columns or set()
    bold = Font(bold=True)
    # Header row.
    for col_ix, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_ix, value=name)
        cell.font = bold
    # Data rows.
    for row_ix, values in enumerate(rows, start=2):
        for col_ix, (name, value) in enumerate(zip(columns, values), start=1):
            cell = ws.cell(row=row_ix, column=col_ix, value=value)
            if name in cost_columns:
                cell.number_format = "$0.00"
            elif name in rate_columns:
                cell.number_format = "0.0%"
            elif isinstance(value, int):
                cell.number_format = "0"
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Self-check
# ---------------------------------------------------------------------------

_SELF_CHECK_FIXTURES = (
    # v2 Phase 2 Step 1: fixture dir names carry the mode segment
    # (`T1-doc-edit-mock`, `T1-doc-edit-real`, `T3-out-of-scope-real`).
    # Dir name == run_id, and the events.jsonl carries that run_id
    # verbatim — the JOIN between events and run_metadata stays on
    # run_id. The T1 mock+real pair gives the self-check direct coverage
    # of the new composite-key invariant: ingesting both does not
    # clobber the mock half.
    #
    # v2 Phase 2 Step 2: T2-two-step-real synthesises the
    # "validation.fail → retry → escalate" episode shape so the
    # `validation_failure_episodes` view has direct fixture coverage.
    # See the fixture's events.jsonl for the inline event sequence.
    "T1-doc-edit-mock",
    "T1-doc-edit-real",
    "T2-two-step-real",
    "T3-out-of-scope-real",
)


def self_check() -> int:
    """Run the harness end-to-end against the two bundled fixtures.

    Returns 0 on PASS, 1 on FAIL. Uses a tmp DuckDB file so the live
    default DB (`state/v2-phase-2/calibration.duckdb`) is untouched.
    `--self-check` is exempt from `--db-path`: it is always hermetic.
    """
    fixture_root = Path(__file__).resolve().parent / "fixtures" / "v2-phase-1"
    try:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            db_file = tmp_path / "self-check.duckdb"
            con = open_db(db_file)

            ingested: list[dict] = []
            for name in _SELF_CHECK_FIXTURES:
                run_dir = fixture_root / name
                if not (run_dir / "events.jsonl").is_file():
                    print(f"self-check: FAIL — missing fixture {run_dir}")
                    return 1
                ingested.append(ingest(con, run_dir))

            # Assertions.
            ops = query_operations(con)
            if len(ops) < 2:
                print(f"self-check: FAIL — operations rows too few ({len(ops)})")
                return 1
            runs = query_per_run_summary(con)
            expected_runs = len(_SELF_CHECK_FIXTURES)
            if len(runs) != expected_runs:
                print(
                    f"self-check: FAIL — per_run_summary expected "
                    f"{expected_runs}, got {len(runs)}"
                )
                return 1
            tasks = query_per_task_comparison(con)
            # v2 Phase 2 Step 1: per_task_comparison groups by task_id;
            # T1 contributes one row (mock + real on a single task_id),
            # T3 contributes one row (real-only). Two task_id rows total.
            if len(tasks) < 1:
                print("self-check: FAIL — per_task_comparison returned 0 rows")
                return 1
            # v2 Phase 2 Step 1: the T1 row must have BOTH mock and real
            # halves populated — the regression that the composite-key
            # fix is meant to prevent.
            t1_row = next(
                (r for r in tasks if r[0] == "T1"), None,
            )
            if t1_row is None:
                print("self-check: FAIL — T1 missing from per_task_comparison")
                return 1
            # Columns per _PER_TASK_COMPARISON_COLUMNS:
            #   (task_id, planner_calls_mock, planner_calls_real,
            #    total_duration_mock, total_duration_real,
            #    total_cost_real, framework_overhead_s)
            if t1_row[1] is None or t1_row[2] is None:
                print(
                    "self-check: FAIL — T1 mock+real key-shape regression: "
                    f"mock={t1_row[1]} real={t1_row[2]}"
                )
                return 1

            # v2 Phase 2 Step 2: validation_failure_episodes must
            # surface the T2-two-step-real fixture's synthetic episode
            # — one row with n_validation_fails=1 and extra_api_calls=1
            # (the retry's stage_b.api_end). Direct regression coverage
            # for the view's two load-bearing aggregates.
            episodes = query_validation_failure_episodes(con)
            t2_episodes = [
                r for r in episodes if r[0] == "T2-two-step-real"
            ]
            if len(t2_episodes) != 1:
                print(
                    "self-check: FAIL — validation_failure_episodes for "
                    f"T2-two-step-real expected 1 row, got {len(t2_episodes)}"
                )
                return 1
            ep = t2_episodes[0]
            # Columns per _VALIDATION_FAILURE_EPISODES_COLUMNS:
            #   (run_id, task_id, mode, step_idx,
            #    n_validation_fails, first_error, second_error,
            #    recovered, extra_api_calls, total_extra_cost_usd)
            if ep[4] != 1:
                print(
                    "self-check: FAIL — T2 episode n_validation_fails "
                    f"expected 1, got {ep[4]}"
                )
                return 1
            if ep[8] != 1:
                print(
                    "self-check: FAIL — T2 episode extra_api_calls "
                    f"expected 1, got {ep[8]}"
                )
                return 1

            # v3 Phase 0 Step 2 (V3P0-3): shadow-decision recorder. The
            # T1 mock+real fixtures each carry 2 paired shadow.decision
            # events (one per planner stage api_end) → 4 shadow rows. The
            # T2/T3 fixtures predate Step 2 and carry no shadow events, so
            # the 1:1 invariant is checked scoped to T1 here; the global
            # "every planner call emits a paired shadow" invariant is
            # verified by the mock-only smoke (fresh full-instrumentation
            # run). policy_version on champion_challenger is NULL in this
            # self-check because the pre-Step-1 golden fixtures' api_end
            # rows lack features_seen — the column populates in the smoke.
            shadow_rows = con.execute(
                "SELECT stage, agreement FROM shadow_decisions"
            ).fetchall()
            t1_planner_api_ends = con.execute(
                "SELECT COUNT(*) FROM events "
                "WHERE kind LIKE 'planner.stage_%.api_end' "
                "  AND run_id IN ('T1-doc-edit-mock', 'T1-doc-edit-real')"
            ).fetchone()[0]
            if len(shadow_rows) != t1_planner_api_ends:
                print(
                    "self-check: FAIL — shadow_decisions rows "
                    f"({len(shadow_rows)}) != T1 planner stage api_end "
                    f"events ({t1_planner_api_ends})"
                )
                return 1
            if not shadow_rows:
                print("self-check: FAIL — no shadow_decisions rows ingested")
                return 1
            # agreement = TRUE on 100% of rows (Phase 0 placeholder always
            # picks Opus, matching the actual Opus route).
            disagreements = [r for r in shadow_rows if r[1] is not True]
            if disagreements:
                print(
                    "self-check: FAIL — shadow agreement not 100% "
                    f"({len(disagreements)} of {len(shadow_rows)} disagree)"
                )
                return 1
            # champion_challenger_comparison returns rows (one per
            # stage × task; T1 contributes stage A and B).
            cc_rows = con.execute(
                "SELECT * FROM champion_challenger_comparison"
            ).fetchall()
            if len(cc_rows) < 1:
                print(
                    "self-check: FAIL — champion_challenger_comparison "
                    "returned 0 rows"
                )
                return 1

            # v3 Phase 0 Step 3 (V3P0-4 / V3P0-5): silent-miss comparator
            # + parser_drop. The T1 fixtures carry shadow_compare.begin/end
            # pairs (one per Stage A api_end) and one synthetic parser_drop.
            # (1) comparator events present and well-formed: every
            #     shadow_compare.end has silent_miss_count=0 + jaccard=1.0
            #     (routed == baseline by construction).
            compare_ends = con.execute(
                "SELECT json_extract(data, '$.silent_miss_count'), "
                "       json_extract(data, '$.jaccard_similarity') "
                "FROM events WHERE kind = 'stage_a.shadow_compare.end'"
            ).fetchall()
            if not compare_ends:
                print("self-check: FAIL — no stage_a.shadow_compare.end events")
                return 1
            for sm, jac in compare_ends:
                if int(sm) != 0 or float(jac) != 1.0:
                    print(
                        "self-check: FAIL — shadow_compare.end not identity "
                        f"(silent_miss={sm}, jaccard={jac})"
                    )
                    return 1
            # (2) silent_miss_episodes view queryable + empty in Phase 0.
            episodes_sm = con.execute(
                "SELECT COUNT(*) FROM silent_miss_episodes"
            ).fetchone()[0]
            if episodes_sm != 0:
                print(
                    "self-check: FAIL — silent_miss_episodes expected 0 rows "
                    f"in Phase 0, got {episodes_sm}"
                )
                return 1
            # (3) parser_drop event present (the fixture's synthetic drop).
            drops = con.execute(
                "SELECT COUNT(*) FROM events WHERE kind = 'stage_a.parser_drop'"
            ).fetchone()[0]
            if drops < 1:
                print("self-check: FAIL — no stage_a.parser_drop events")
                return 1

            # v3 Phase 0 Step 4 (V3P0-6): cache-family diagnostics. The T1
            # fixtures carry the three new fields on their planner stage
            # events; the cache_diagnostics view exposes them per-stage.
            cache_rows = con.execute(
                "SELECT COUNT(*) FROM cache_diagnostics"
            ).fetchone()[0]
            if cache_rows < 1:
                print("self-check: FAIL — cache_diagnostics returned 0 rows")
                return 1
            # Stage A on the T1 mock fixture: vault_index_hit populated
            # (bool, not null — the question applies to Stage A) and the
            # candidate block decomposition sums to a positive estimate.
            stage_a = con.execute(
                "SELECT vault_index_hit, candidate_block_sizes_sum "
                "FROM cache_diagnostics "
                "WHERE run_id = 'T1-doc-edit-mock' AND stage = 'A'"
            ).fetchone()
            if stage_a is None:
                print("self-check: FAIL — no Stage A cache_diagnostics row for T1")
                return 1
            if stage_a[0] is None:
                print("self-check: FAIL — Stage A vault_index_hit is null "
                      "(should be a bool — the field applies to Stage A)")
                return 1
            if (stage_a[1] or 0) <= 0:
                print("self-check: FAIL — Stage A candidate_block_sizes_sum "
                      f"non-positive ({stage_a[1]})")
                return 1
            # Stage B vault_index_hit is null by construction (Q(c)).
            stage_b = con.execute(
                "SELECT vault_index_hit FROM cache_diagnostics "
                "WHERE run_id = 'T1-doc-edit-mock' AND stage = 'B'"
            ).fetchone()
            if stage_b is not None and stage_b[0] is not None:
                print("self-check: FAIL — Stage B vault_index_hit should be "
                      f"null (n/a), got {stage_b[0]}")
                return 1

            xlsx_path = tmp_path / "self-check.xlsx"
            export_xlsx(con, xlsx_path)
            if not xlsx_path.is_file():
                print(f"self-check: FAIL — XLSX not created at {xlsx_path}")
                return 1
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path)
            expected_sheets = {
                "operations",
                "per_run_summary",
                "per_task_comparison",
                # v2 Phase 2 Step 2: validation_episodes sheet is the
                # XLSX surface of the new view.
                "validation_episodes",
            }
            actual_sheets = set(wb.sheetnames)
            if not expected_sheets.issubset(actual_sheets):
                missing = expected_sheets - actual_sheets
                print(f"self-check: FAIL — XLSX missing sheets {missing}")
                return 1
            # Header row bold check.
            ws = wb["operations"]
            if not ws.cell(row=1, column=1).font.bold:
                print("self-check: FAIL — operations header row not bold")
                return 1

            con.close()
        print(
            f"self-check: PASS ({len(ops)} ops, {len(runs)} runs, "
            f"{len(tasks)} tasks, {len(episodes)} episodes, "
            f"{len(shadow_rows)} shadow, {len(compare_ends)} compare, "
            f"{drops} parser_drop, {cache_rows} cache_diag)"
        )
        return 0
    except Exception as e:  # noqa: BLE001
        print(f"self-check: FAIL — {type(e).__name__}: {e}")
        return 1


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _print_rows(columns: tuple[str, ...], rows: Iterable[tuple]) -> None:
    """Compact tab-separated stdout output. Suited for grep / paste into
    spreadsheets; not for human-pretty-printing wide tables."""
    sys.stdout.write("\t".join(columns) + "\n")
    for r in rows:
        sys.stdout.write("\t".join("" if v is None else str(v) for v in r) + "\n")


def _resolve_db_path(cli_db_path: str | None) -> Path:
    """v2 Phase 3 Step 1: resolve the CLI DuckDB path. An explicit
    `--db-path` wins (expanded, but not forced-absolute so relative
    paths still resolve from cwd); otherwise the `db_path()` default
    (v2-phase-2)."""
    if cli_db_path:
        return Path(cli_db_path).expanduser()
    return db_path()


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        prog="harness_v2",
        description="ANVIL v2 Phase 1 event-stream harness.",
    )
    parser.add_argument("--self-check", action="store_true",
                        help="Run against bundled fixtures; print PASS/FAIL.")
    # v2 Phase 3 Step 1: top-level --db-path mirrors calibration_runner's
    # flag (Finding 1). Threaded into the single open_db() below, so every
    # DB-touching subcommand honours it. Must precede the subcommand on the
    # command line (top-level arg under argparse subparsers). Default
    # resolves to db_path() (v2-phase-2 — the clean baseline). --self-check
    # is exempt: it stays hermetic against its own tmp DB.
    parser.add_argument(
        "--db-path", type=str, default=None,
        help=(
            "DuckDB path for all query/ingest subcommands. "
            "Default: <ANVIL_ROOT>/state/v2-phase-2/calibration.duckdb"
        ),
    )
    sub = parser.add_subparsers(dest="command")

    p_ing = sub.add_parser("ingest", help="Ingest one run-dir.")
    p_ing.add_argument("run_dir", type=Path)

    p_ingall = sub.add_parser("ingest-all", help="Ingest every state/runs/* dir.")
    p_ingall.add_argument("--state-root", type=Path, default=None)

    p_ops = sub.add_parser("operations", help="Print operations view.")
    # v2 Phase 5 Step 1b: positional run_id, mirroring per-run-summary
    # (Finding 4 — Tier 4 CLI consistency). `operations T1-doc-edit-real`
    # now works like `per-run-summary T1-doc-edit-real`. Replaces the prior
    # --run-id flag (consumed by no code/test — only the module docstring).
    p_ops.add_argument("run_id", nargs="?", default=None,
                       help="optional run_id; omit for all runs")

    p_runs = sub.add_parser("per-run-summary", help="Print per_run_summary view.")
    # v2 Phase 3 Step 1: run_id is a positional (optional) so the brief's
    # smoke command shape works: `per-run-summary T1-doc-edit-real`.
    # Omit for all runs. (Replaces the prior --run-id flag on this
    # subcommand; no code/test consumed the flag form.)
    p_runs.add_argument("run_id", nargs="?", default=None,
                        help="optional run_id; omit for all runs")

    sub.add_parser("per-task-comparison", help="Print per_task_comparison view.")

    # v2 Phase 2 Step 2: validation_failure_episodes is the v2 Phase 2
    # exam Q2 view. Same shape as the other view sub-commands: prints
    # tab-separated rows; pipe to less / a spreadsheet / grep.
    sub.add_parser(
        "validation-failure-episodes",
        help="Print validation_failure_episodes view.",
    )

    p_xlsx = sub.add_parser(
        "export-xlsx",
        help="Export four sheets to <out.xlsx>.",
    )
    p_xlsx.add_argument("out_path", type=Path)

    # v2 Phase 3 Step 1: cumulative-spend prints cumulative_spend_by_task
    # (actual-spend reconciliation). Optional --mode filter (real/mock).
    p_cum = sub.add_parser(
        "cumulative-spend",
        help="Print cumulative_spend_by_task view (actual-spend reconciliation).",
    )
    p_cum.add_argument("--mode", type=str, default=None,
                       choices=["real", "mock", "unknown"],
                       help="optional mode filter")

    args = parser.parse_args(argv)

    if args.self_check:
        return self_check()

    if args.command is None:
        parser.print_help()
        return 2

    con = open_db(_resolve_db_path(args.db_path))
    try:
        if args.command == "ingest":
            summary = ingest(con, args.run_dir)
            print(json.dumps(summary, indent=2))
            return 0
        if args.command == "ingest-all":
            summaries = ingest_all(con, args.state_root)
            print(json.dumps(summaries, indent=2))
            return 0
        if args.command == "operations":
            _print_rows(_OPERATIONS_COLUMNS, query_operations(con, args.run_id))
            return 0
        if args.command == "per-run-summary":
            _print_rows(_PER_RUN_SUMMARY_COLUMNS, query_per_run_summary(con, args.run_id))
            return 0
        if args.command == "per-task-comparison":
            _print_rows(_PER_TASK_COMPARISON_COLUMNS, query_per_task_comparison(con))
            return 0
        if args.command == "validation-failure-episodes":
            _print_rows(
                _VALIDATION_FAILURE_EPISODES_COLUMNS,
                query_validation_failure_episodes(con),
            )
            return 0
        if args.command == "export-xlsx":
            export_xlsx(con, args.out_path)
            print(f"wrote {args.out_path}")
            return 0
        if args.command == "cumulative-spend":
            _print_rows(
                _CUMULATIVE_SPEND_COLUMNS,
                query_cumulative_spend(con, args.mode),
            )
            return 0
    finally:
        con.close()
    return 2


if __name__ == "__main__":
    sys.exit(main())
